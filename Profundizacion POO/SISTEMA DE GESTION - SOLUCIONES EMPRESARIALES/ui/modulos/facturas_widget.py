from datetime import date

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QTableWidget,
    QTableWidgetItem, QDialog, QFormLayout, QLineEdit, QComboBox,
    QLabel, QMessageBox, QHeaderView, QFileDialog, QScrollArea,
    QDoubleSpinBox, QDateEdit, QTextEdit
)
from PyQt6.QtCore import Qt, QDate

from models.factura import Factura
from ui.themes import obtener_tema, estilo_boton, estilo_input, estilo_tabla
from ui.iconos.iconos import iconos_crud, iconos_formulario, favicon


class FacturasWidget(QWidget):

    def __init__(self, repo):
        super().__init__()
        self.repo = repo
        self._construir_ui()
        self._aplicar_tema()
        self._cargar_datos()

    def _construir_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(16, 16, 16, 16)
        self.setLayout(layout)

        self.titulo = QLabel("Facturas")
        self.titulo.setStyleSheet("font-size: 22px; font-weight: bold; padding: 4px 0;")
        layout.addWidget(self.titulo)

        barra = QHBoxLayout()
        self.btn_nuevo    = QPushButton("  Nuevo")
        self.btn_editar   = QPushButton("  Editar")
        self.btn_eliminar = QPushButton("  Eliminar")
        self.btn_excel    = QPushButton("  Exportar Excel")
        self.btn_pdf      = QPushButton("  Exportar PDF")
        for btn in [self.btn_nuevo, self.btn_editar, self.btn_eliminar,
                    self.btn_excel, self.btn_pdf]:
            btn.setMinimumHeight(34)
        barra.addWidget(self.btn_nuevo)
        barra.addWidget(self.btn_editar)
        barra.addWidget(self.btn_eliminar)
        barra.addSpacing(20)
        barra.addWidget(self.btn_excel)
        barra.addWidget(self.btn_pdf)
        barra.addStretch()
        layout.addLayout(barra)

        filtros = QHBoxLayout()
        self.campo_busqueda = QLineEdit()
        self.campo_busqueda.setPlaceholderText("Buscar por numero o cliente...")
        self.campo_busqueda.setMinimumHeight(32)

        self.filtro_estado = QComboBox()
        self.filtro_estado.addItems(["Todos los estados", "emitida", "enviada",
                                     "pagada", "vencida", "anulada"])
        self.filtro_estado.setMinimumHeight(32)

        lbl_desde = QLabel("Desde:")
        lbl_desde.setStyleSheet("font-size: 12px;")
        self.filtro_desde = QDateEdit()
        self.filtro_desde.setCalendarPopup(True)
        self.filtro_desde.setDate(QDate(QDate.currentDate().year(), 1, 1))
        self.filtro_desde.setDisplayFormat("dd/MM/yyyy")
        self.filtro_desde.setMinimumHeight(32)

        lbl_hasta = QLabel("Hasta:")
        lbl_hasta.setStyleSheet("font-size: 12px;")
        self.filtro_hasta = QDateEdit()
        self.filtro_hasta.setCalendarPopup(True)
        self.filtro_hasta.setDate(QDate.currentDate())
        self.filtro_hasta.setDisplayFormat("dd/MM/yyyy")
        self.filtro_hasta.setMinimumHeight(32)

        self.btn_buscar  = QPushButton("  Buscar")
        self.btn_limpiar = QPushButton("  Limpiar")
        self.btn_buscar.setMinimumHeight(32)
        self.btn_limpiar.setMinimumHeight(32)

        filtros.addWidget(self.campo_busqueda, 2)
        filtros.addWidget(self.filtro_estado, 1)
        filtros.addWidget(lbl_desde)
        filtros.addWidget(self.filtro_desde)
        filtros.addWidget(lbl_hasta)
        filtros.addWidget(self.filtro_hasta)
        filtros.addWidget(self.btn_buscar)
        filtros.addWidget(self.btn_limpiar)
        layout.addLayout(filtros)

        self.tabla = QTableWidget()
        columnas = ["Numero", "Fecha", "Cliente", "Proyecto",
                    "Honorarios", "Gastos", "Descuentos", "Impuestos",
                    "Total", "Condiciones", "Estado"]
        self.tabla.setColumnCount(len(columnas))
        self.tabla.setHorizontalHeaderLabels(columnas)
        self.tabla.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.tabla.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.tabla.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.tabla.horizontalHeader().setStretchLastSection(True)
        self.tabla.verticalHeader().setVisible(False)
        self.tabla.setAlternatingRowColors(True)
        layout.addWidget(self.tabla)

        self.lbl_total = QLabel("Total: 0 registros  |  Total facturado: $0.00")
        layout.addWidget(self.lbl_total)

        self.btn_nuevo.clicked.connect(self._abrir_formulario_nuevo)
        self.btn_editar.clicked.connect(self._abrir_formulario_editar)
        self.btn_eliminar.clicked.connect(self._eliminar)
        self.btn_excel.clicked.connect(self._exportar_excel)
        self.btn_pdf.clicked.connect(self._exportar_pdf)
        self.btn_buscar.clicked.connect(self._buscar)
        self.btn_limpiar.clicked.connect(self._limpiar)
        self.campo_busqueda.returnPressed.connect(self._buscar)
        self.filtro_estado.currentIndexChanged.connect(self._buscar)

        iconos_crud(self.btn_nuevo, self.btn_editar, self.btn_eliminar,
                    self.btn_excel, self.btn_pdf, self.btn_buscar, self.btn_limpiar)

    def _aplicar_tema(self):
        t = obtener_tema()
        self.setStyleSheet(f"background-color: {t['fondo']}; color: {t['texto']};")
        self.titulo.setStyleSheet(f"font-size: 22px; font-weight: bold; color: {t['titulo']};")
        self.lbl_total.setStyleSheet(f"font-size: 11px; color: {t['texto_suave']};")
        self.btn_nuevo.setStyleSheet(estilo_boton(t["btn_primario"],  t["btn_primario_txt"]))
        self.btn_editar.setStyleSheet(estilo_boton(t["btn_editar"],   t["btn_editar_txt"]))
        self.btn_eliminar.setStyleSheet(estilo_boton(t["btn_peligro"],t["btn_peligro_txt"]))
        self.btn_excel.setStyleSheet(estilo_boton(t["btn_excel"],     t["btn_excel_txt"]))
        self.btn_pdf.setStyleSheet(estilo_boton(t["btn_pdf"],         t["btn_pdf_txt"]))
        self.btn_buscar.setStyleSheet(estilo_boton(t["btn_buscar"],   t["btn_buscar_txt"]))
        self.btn_limpiar.setStyleSheet(estilo_boton(t["btn_limpiar"], t["btn_limpiar_txt"]))
        self.campo_busqueda.setStyleSheet(estilo_input(t))
        self.filtro_estado.setStyleSheet(estilo_input(t))
        self.filtro_desde.setStyleSheet(estilo_input(t))
        self.filtro_hasta.setStyleSheet(estilo_input(t))
        self.tabla.setStyleSheet(estilo_tabla(t))

    def _total_factura(self, f):
        return (float(f.honorarios or 0) + float(f.gastos_reembolsables or 0)
                - float(f.descuentos or 0) + float(f.impuestos or 0))

    def _cargar_datos(self, facturas=None):
        if facturas is None:
            facturas = self.repo.select_all()
        self.tabla.setRowCount(len(facturas))
        total_facturado = 0
        for fila, f in enumerate(facturas):
            total = self._total_factura(f)
            total_facturado += total
            self.tabla.setItem(fila, 0,  QTableWidgetItem(str(f.numero_factura or "")))
            self.tabla.setItem(fila, 1,  QTableWidgetItem(str(f.fecha or "")))
            self.tabla.setItem(fila, 2,  QTableWidgetItem(str(f.cliente_codigo or "")))
            self.tabla.setItem(fila, 3,  QTableWidgetItem(str(f.proyecto_numero or "")))
            self.tabla.setItem(fila, 4,  QTableWidgetItem(f"${float(f.honorarios or 0):,.2f}"))
            self.tabla.setItem(fila, 5,  QTableWidgetItem(f"${float(f.gastos_reembolsables or 0):,.2f}"))
            self.tabla.setItem(fila, 6,  QTableWidgetItem(f"${float(f.descuentos or 0):,.2f}"))
            self.tabla.setItem(fila, 7,  QTableWidgetItem(f"${float(f.impuestos or 0):,.2f}"))
            self.tabla.setItem(fila, 8,  QTableWidgetItem(f"${total:,.2f}"))
            self.tabla.setItem(fila, 9,  QTableWidgetItem(str(f.condiciones_pago or "")))
            self.tabla.setItem(fila, 10, QTableWidgetItem(str(f.estado or "")))
        self.tabla.resizeColumnsToContents()
        self.lbl_total.setText(
            f"Total: {len(facturas)} registros  |  Total facturado: ${total_facturado:,.2f}"
        )

    def _fila_seleccionada(self):
        fila = self.tabla.currentRow()
        if fila == -1:
            return None
        return self.tabla.item(fila, 0).text()

    def _buscar(self):
        termino = self.campo_busqueda.text().strip().lower()
        estado  = self.filtro_estado.currentText()
        desde_q = self.filtro_desde.date()
        hasta_q = self.filtro_hasta.date()
        desde   = date(desde_q.year(), desde_q.month(), desde_q.day())
        hasta   = date(hasta_q.year(), hasta_q.month(), hasta_q.day())

        resultados = self.repo.select_all()

        if termino:
            resultados = [f for f in resultados
                          if termino in str(f.numero_factura or "").lower()
                          or termino in str(f.cliente_codigo or "").lower()]
        if estado != "Todos los estados":
            resultados = [f for f in resultados if f.estado == estado]

        resultados = [f for f in resultados
                      if f.fecha and desde <= f.fecha <= hasta]

        self._cargar_datos(resultados)

    def _limpiar(self):
        self.campo_busqueda.clear()
        self.filtro_estado.setCurrentIndex(0)
        self.filtro_desde.setDate(QDate(QDate.currentDate().year(), 1, 1))
        self.filtro_hasta.setDate(QDate.currentDate())
        self._cargar_datos()

    def _abrir_formulario_nuevo(self):
        dialogo = FormularioFactura(self.repo)
        if dialogo.exec() == QDialog.DialogCode.Accepted:
            self._cargar_datos()

    def _abrir_formulario_editar(self):
        numero = self._fila_seleccionada()
        if not numero:
            QMessageBox.warning(self, "Aviso", "Selecciona una factura para editar.")
            return
        factura = self.repo.select_WHERE_numero_factura(numero)
        dialogo = FormularioFactura(self.repo, factura)
        if dialogo.exec() == QDialog.DialogCode.Accepted:
            self._cargar_datos()

    def _eliminar(self):
        numero = self._fila_seleccionada()
        if not numero:
            QMessageBox.warning(self, "Aviso", "Selecciona una factura para eliminar.")
            return
        respuesta = QMessageBox.question(
            self, "Confirmar eliminacion",
            f"Estas seguro de eliminar la factura {numero}?\n\nEsta accion no se puede deshacer.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if respuesta == QMessageBox.StandardButton.Yes:
            try:
                self.repo.delete_factura(numero)
                self._cargar_datos()
                QMessageBox.information(self, "Exito", "Factura eliminada correctamente.")
            except Exception as e:
                QMessageBox.critical(self, "No se puede eliminar", str(e))

    def _exportar_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            QMessageBox.critical(self, "Error", "Instala openpyxl:\npip install openpyxl")
            return

        dialogo = DialogoFiltroFacturas(self)
        if dialogo.exec() != QDialog.DialogCode.Accepted:
            return

        datos   = self._aplicar_filtros(dialogo.obtener_filtros())
        ruta, _ = QFileDialog.getSaveFileName(self, "Guardar Excel", "facturas.xlsx", "Excel (*.xlsx)")
        if not ruta:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Facturas"
        fill_enc = PatternFill("solid", fgColor="14532D")
        font_enc = Font(bold=True, color="FFFFFF", size=11)
        font_tit = Font(bold=True, size=14, color="14532D")
        borde    = Border(left=Side(style="thin"), right=Side(style="thin"),
                          top=Side(style="thin"),  bottom=Side(style="thin"))
        fill_par = PatternFill("solid", fgColor="DCFCE7")

        ws.merge_cells("A1:K1")
        ws["A1"] = "Reporte de Facturas"
        ws["A1"].font      = font_tit
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:K2")
        ws["A2"] = f"Generado: {date.today().strftime('%d/%m/%Y')}  |  Total: {len(datos)} registros"
        ws["A2"].alignment = Alignment(horizontal="center")
        ws["A2"].font      = Font(italic=True, color="666666")

        cols = ["Numero", "Fecha", "Cliente", "Proyecto", "Honorarios",
                "Gastos", "Descuentos", "Impuestos", "Total", "Condiciones", "Estado"]
        for i, col in enumerate(cols, 1):
            cell           = ws.cell(row=4, column=i, value=col)
            cell.font      = font_enc
            cell.fill      = fill_enc
            cell.alignment = Alignment(horizontal="center")
            cell.border    = borde

        for idx, f in enumerate(datos, 5):
            total     = self._total_factura(f)
            fila_data = [f.numero_factura, str(f.fecha or ""), f.cliente_codigo,
                         f.proyecto_numero, float(f.honorarios or 0),
                         float(f.gastos_reembolsables or 0), float(f.descuentos or 0),
                         float(f.impuestos or 0), total, f.condiciones_pago, f.estado]
            fill = fill_par if idx % 2 == 0 else None
            for j, val in enumerate(fila_data, 1):
                cell        = ws.cell(row=idx, column=j, value=val or "")
                cell.border = borde
                if fill:
                    cell.fill = fill

        anchos = [18, 12, 14, 14, 14, 12, 12, 12, 14, 20, 12]
        for i, ancho in enumerate(anchos, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

        wb.save(ruta)
        QMessageBox.information(self, "Exportado", f"Excel guardado en:\n{ruta}")

    def _exportar_pdf(self):
        try:
            from reportlab.lib.pagesizes import landscape, A3
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import ParagraphStyle
            from reportlab.lib.units import cm
        except ImportError:
            QMessageBox.critical(self, "Error", "Instala reportlab:\npip install reportlab")
            return

        dialogo = DialogoFiltroFacturas(self)
        if dialogo.exec() != QDialog.DialogCode.Accepted:
            return

        datos   = self._aplicar_filtros(dialogo.obtener_filtros())
        ruta, _ = QFileDialog.getSaveFileName(self, "Guardar PDF", "facturas.pdf", "PDF (*.pdf)")
        if not ruta:
            return

        doc   = SimpleDocTemplate(ruta, pagesize=landscape(A3),
                                  topMargin=1.5*cm, bottomMargin=1.5*cm)
        story = []
        st = ParagraphStyle("t", fontSize=16, fontName="Helvetica-Bold",
                             spaceAfter=4, textColor=colors.HexColor("#14532D"))
        ss = ParagraphStyle("s", fontSize=9, fontName="Helvetica",
                             spaceAfter=12, textColor=colors.grey)
        total_general = sum(self._total_factura(f) for f in datos)
        story.append(Paragraph("Reporte de Facturas", st))
        story.append(Paragraph(
            f"Generado: {date.today().strftime('%d/%m/%Y')}  |  "
            f"Total registros: {len(datos)}  |  "
            f"Total facturado: ${total_general:,.2f}", ss))
        story.append(Spacer(1, 0.3*cm))

        enc   = ["Numero", "Fecha", "Cliente", "Proyecto",
                 "Honorarios", "Total", "Estado"]
        filas = [enc]
        for f in datos:
            filas.append([
                f.numero_factura or "", str(f.fecha or ""),
                f.cliente_codigo or "", f.proyecto_numero or "",
                f"${float(f.honorarios or 0):,.2f}",
                f"${self._total_factura(f):,.2f}",
                f.estado or ""
            ])

        tabla = Table(filas, repeatRows=1)
        tabla.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),(-1,0),  colors.HexColor("#14532D")),
            ("TEXTCOLOR",     (0,0),(-1,0),  colors.white),
            ("FONTNAME",      (0,0),(-1,0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0,0),(-1,-1), 8),
            ("ALIGN",         (0,0),(-1,-1), "CENTER"),
            ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, colors.HexColor("#DCFCE7")]),
            ("GRID",          (0,0),(-1,-1), 0.5, colors.HexColor("#CCCCCC")),
            ("TOPPADDING",    (0,0),(-1,-1), 5),
            ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ]))
        story.append(tabla)
        doc.build(story)
        QMessageBox.information(self, "Exportado", f"PDF guardado en:\n{ruta}")

    def _aplicar_filtros(self, filtros):
        estado = filtros.get("estado")
        desde  = filtros.get("desde")
        hasta  = filtros.get("hasta")
        datos  = self.repo.select_all()
        if estado:
            datos = [f for f in datos if f.estado == estado]
        if desde and hasta:
            datos = [f for f in datos if f.fecha and desde <= f.fecha <= hasta]
        return datos


class DialogoFiltroFacturas(QDialog):

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Filtros de exportacion")
        self.setWindowIcon(favicon("facturas"))
        self.setMinimumWidth(340)
        t = obtener_tema()
        self.setStyleSheet(f"background-color: {t['fondo_panel']}; color: {t['texto']};")
        layout = QVBoxLayout()
        self.setLayout(layout)

        layout.addWidget(QLabel("Filtrar por estado:"))
        self.combo_estado = QComboBox()
        self.combo_estado.addItems(["Todos", "emitida", "enviada", "pagada", "vencida", "anulada"])
        self.combo_estado.setStyleSheet(estilo_input(t))
        layout.addWidget(self.combo_estado)

        layout.addWidget(QLabel("Desde:"))
        self.desde = QDateEdit()
        self.desde.setCalendarPopup(True)
        self.desde.setDate(QDate(QDate.currentDate().year(), 1, 1))
        self.desde.setDisplayFormat("dd/MM/yyyy")
        self.desde.setStyleSheet(estilo_input(t))
        layout.addWidget(self.desde)

        layout.addWidget(QLabel("Hasta:"))
        self.hasta = QDateEdit()
        self.hasta.setCalendarPopup(True)
        self.hasta.setDate(QDate.currentDate())
        self.hasta.setDisplayFormat("dd/MM/yyyy")
        self.hasta.setStyleSheet(estilo_input(t))
        layout.addWidget(self.hasta)

        barra      = QHBoxLayout()
        btn_ok     = QPushButton("  Exportar")
        btn_cancel = QPushButton("  Cancelar")
        btn_ok.setStyleSheet(estilo_boton(t["btn_primario"], t["btn_primario_txt"]))
        btn_cancel.setStyleSheet(estilo_boton(t["btn_limpiar"], t["btn_limpiar_txt"]))
        barra.addStretch()
        barra.addWidget(btn_ok)
        barra.addWidget(btn_cancel)
        layout.addLayout(barra)
        btn_ok.clicked.connect(self.accept)
        btn_cancel.clicked.connect(self.reject)

    def obtener_filtros(self):
        estado  = self.combo_estado.currentText()
        desde_q = self.desde.date()
        hasta_q = self.hasta.date()
        return {
            "estado": None if estado == "Todos" else estado,
            "desde":  date(desde_q.year(), desde_q.month(), desde_q.day()),
            "hasta":  date(hasta_q.year(), hasta_q.month(), hasta_q.day()),
        }


class FormularioFactura(QDialog):

    def __init__(self, repo, factura=None):
        super().__init__()
        self.repo     = repo
        self.factura  = factura
        self.editando = factura is not None
        t             = obtener_tema()

        self.setWindowTitle("Editar Factura" if self.editando else "Nueva Factura")
        self.setWindowIcon(favicon("facturas"))
        self.setMinimumWidth(480)
        self.setStyleSheet(f"background-color: {t['fondo_panel']}; color: {t['texto']};")

        layout = QVBoxLayout()
        self.setLayout(layout)

        scroll     = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("border: none;")
        contenedor = QWidget()
        scroll.setWidget(contenedor)
        form_layout = QVBoxLayout()
        contenedor.setLayout(form_layout)

        form   = QFormLayout()
        form.setSpacing(10)
        estilo = estilo_input(t)

        self.campo_numero    = QLineEdit()
        self.campo_fecha     = QDateEdit()
        self.campo_fecha.setCalendarPopup(True)
        self.campo_fecha.setDate(QDate.currentDate())
        self.campo_fecha.setDisplayFormat("dd/MM/yyyy")
        self.campo_cliente   = QLineEdit()
        self.campo_proyecto  = QLineEdit()
        self.campo_p_inicio  = QDateEdit()
        self.campo_p_inicio.setCalendarPopup(True)
        self.campo_p_inicio.setDate(QDate.currentDate())
        self.campo_p_inicio.setDisplayFormat("dd/MM/yyyy")
        self.campo_p_fin     = QDateEdit()
        self.campo_p_fin.setCalendarPopup(True)
        self.campo_p_fin.setDate(QDate.currentDate())
        self.campo_p_fin.setDisplayFormat("dd/MM/yyyy")
        self.campo_servicios = QTextEdit()
        self.campo_servicios.setMaximumHeight(60)
        self.campo_honorarios = QDoubleSpinBox()
        self.campo_honorarios.setRange(0, 99999999.99)
        self.campo_honorarios.setDecimals(2)
        self.campo_honorarios.setPrefix("$ ")
        self.campo_gastos    = QDoubleSpinBox()
        self.campo_gastos.setRange(0, 99999999.99)
        self.campo_gastos.setDecimals(2)
        self.campo_gastos.setPrefix("$ ")
        self.campo_descuentos = QDoubleSpinBox()
        self.campo_descuentos.setRange(0, 99999999.99)
        self.campo_descuentos.setDecimals(2)
        self.campo_descuentos.setPrefix("$ ")
        self.campo_impuestos = QDoubleSpinBox()
        self.campo_impuestos.setRange(0, 99999999.99)
        self.campo_impuestos.setDecimals(2)
        self.campo_impuestos.setPrefix("$ ")
        self.campo_condiciones = QLineEdit()
        self.campo_estado    = QComboBox()
        self.campo_estado.addItems(["emitida", "enviada", "pagada", "vencida", "anulada"])

        for w in [self.campo_numero, self.campo_fecha, self.campo_cliente,
                  self.campo_proyecto, self.campo_p_inicio, self.campo_p_fin,
                  self.campo_servicios, self.campo_honorarios, self.campo_gastos,
                  self.campo_descuentos, self.campo_impuestos,
                  self.campo_condiciones, self.campo_estado]:
            w.setStyleSheet(estilo)

        form.addRow("Numero *:",        self.campo_numero)
        form.addRow("Fecha *:",         self.campo_fecha)
        form.addRow("Cliente *:",       self.campo_cliente)
        form.addRow("Proyecto:",        self.campo_proyecto)
        form.addRow("Periodo Inicio:",  self.campo_p_inicio)
        form.addRow("Periodo Fin:",     self.campo_p_fin)
        form.addRow("Servicios:",       self.campo_servicios)
        form.addRow("Honorarios:",      self.campo_honorarios)
        form.addRow("Gastos:",          self.campo_gastos)
        form.addRow("Descuentos:",      self.campo_descuentos)
        form.addRow("Impuestos:",       self.campo_impuestos)
        form.addRow("Condiciones:",     self.campo_condiciones)
        form.addRow("Estado *:",        self.campo_estado)
        form_layout.addLayout(form)

        barra = QHBoxLayout()
        self.btn_guardar  = QPushButton("  Guardar")
        self.btn_cancelar = QPushButton("  Cancelar")
        self.btn_guardar.setMinimumHeight(34)
        self.btn_cancelar.setMinimumHeight(34)
        self.btn_guardar.setStyleSheet(estilo_boton(t["btn_primario"],  t["btn_primario_txt"]))
        self.btn_cancelar.setStyleSheet(estilo_boton(t["btn_limpiar"],  t["btn_limpiar_txt"]))
        barra.addStretch()
        barra.addWidget(self.btn_guardar)
        barra.addWidget(self.btn_cancelar)
        form_layout.addLayout(barra)

        self.btn_guardar.clicked.connect(self._guardar)
        self.btn_cancelar.clicked.connect(self.reject)
        layout.addWidget(scroll)

        iconos_formulario(self.btn_guardar, self.btn_cancelar)

        if self.editando:
            self.campo_numero.setText(factura.numero_factura)
            self.campo_numero.setEnabled(False)
            if factura.fecha:
                f = factura.fecha
                self.campo_fecha.setDate(QDate(f.year, f.month, f.day))
            self.campo_cliente.setText(factura.cliente_codigo or "")
            self.campo_proyecto.setText(factura.proyecto_numero or "")
            if factura.periodo_inicio:
                p = factura.periodo_inicio
                self.campo_p_inicio.setDate(QDate(p.year, p.month, p.day))
            if factura.periodo_fin:
                p = factura.periodo_fin
                self.campo_p_fin.setDate(QDate(p.year, p.month, p.day))
            self.campo_servicios.setPlainText(factura.servicios_prestados or "")
            self.campo_honorarios.setValue(float(factura.honorarios or 0))
            self.campo_gastos.setValue(float(factura.gastos_reembolsables or 0))
            self.campo_descuentos.setValue(float(factura.descuentos or 0))
            self.campo_impuestos.setValue(float(factura.impuestos or 0))
            self.campo_condiciones.setText(factura.condiciones_pago or "")
            self.campo_estado.setCurrentText(factura.estado or "emitida")

    def _validar(self):
        if not self.campo_numero.text().strip():
            QMessageBox.warning(self, "Error", "El numero de factura es obligatorio.")
            return False
        if len(self.campo_numero.text().strip()) < 3:
            QMessageBox.warning(self, "Error", "El numero debe tener al menos 3 caracteres.")
            return False
        if not self.campo_cliente.text().strip():
            QMessageBox.warning(self, "Error", "El codigo de cliente es obligatorio.")
            return False
        if self.campo_honorarios.value() < 0:
            QMessageBox.warning(self, "Error", "Los honorarios no pueden ser negativos.")
            return False
        return True

    def _guardar(self):
        if not self._validar():
            return
        accion = "actualizar" if self.editando else "guardar"
        respuesta = QMessageBox.question(
            self, "Confirmar",
            f"Deseas {accion} esta factura?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if respuesta != QMessageBox.StandardButton.Yes:
            return

        def qdate_to_date(qd):
            return date(qd.year(), qd.month(), qd.day())

        factura = Factura(
            numero_factura       = self.campo_numero.text().strip(),
            fecha                = qdate_to_date(self.campo_fecha.date()),
            cliente_codigo       = self.campo_cliente.text().strip(),
            proyecto_numero      = self.campo_proyecto.text().strip() or None,
            periodo_inicio       = qdate_to_date(self.campo_p_inicio.date()),
            periodo_fin          = qdate_to_date(self.campo_p_fin.date()),
            servicios_prestados  = self.campo_servicios.toPlainText().strip() or None,
            honorarios           = self.campo_honorarios.value(),
            gastos_reembolsables = self.campo_gastos.value(),
            descuentos           = self.campo_descuentos.value(),
            impuestos            = self.campo_impuestos.value(),
            condiciones_pago     = self.campo_condiciones.text().strip() or None,
            estado               = self.campo_estado.currentText()
        )

        try:
            if self.editando:
                self.repo.update(factura)
            else:
                self.repo.insert(factura)
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Error al guardar", str(e))

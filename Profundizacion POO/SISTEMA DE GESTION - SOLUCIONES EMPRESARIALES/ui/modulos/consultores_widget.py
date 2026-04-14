import os
import re
import shutil
from datetime import date

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QTableWidget,
    QTableWidgetItem, QDialog, QFormLayout, QLineEdit, QComboBox,
    QLabel, QMessageBox, QHeaderView, QDateEdit, QFileDialog,
    QScrollArea, QFrame, QSpinBox, QDoubleSpinBox
)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QPixmap

from models.consultor import Consultor
from ui.themes import obtener_tema, estilo_boton, estilo_input, estilo_tabla
from ui.iconos.iconos import iconos_crud, iconos_formulario, favicon

CARPETA_IMAGENES = os.path.join(os.path.dirname(__file__), "..", "..", "assets", "consultores")
os.makedirs(CARPETA_IMAGENES, exist_ok=True)


class ConsultoresWidget(QWidget):

    def __init__(self, repo):
        super().__init__()
        self.repo = repo
        self._construir_ui()
        self._aplicar_tema()
        self._cargar_datos()

    def _construir_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(16, 16, 16, 16)
        self.setLayout(layout)

        self.titulo = QLabel("Consultores")
        self.titulo.setStyleSheet("font-size: 22px; font-weight: bold; padding: 4px 0;")
        layout.addWidget(self.titulo)

        barra = QHBoxLayout()
        self.btn_nuevo    = QPushButton("  Nuevo")
        self.btn_editar   = QPushButton("  Editar")
        self.btn_eliminar = QPushButton("  Eliminar")
        self.btn_excel    = QPushButton("  Exportar Excel")
        self.btn_pdf      = QPushButton("  Exportar PDF")
        for btn in [self.btn_nuevo, self.btn_editar, self.btn_eliminar,
                    self.btn_excel, self.btn_pdf]:
            btn.setMinimumHeight(34)
        barra.addWidget(self.btn_nuevo)
        barra.addWidget(self.btn_editar)
        barra.addWidget(self.btn_eliminar)
        barra.addSpacing(20)
        barra.addWidget(self.btn_excel)
        barra.addWidget(self.btn_pdf)
        barra.addStretch()
        layout.addLayout(barra)

        filtros = QHBoxLayout()
        self.campo_busqueda = QLineEdit()
        self.campo_busqueda.setPlaceholderText("Buscar por nombre o especialidad...")
        self.campo_busqueda.setMinimumHeight(32)
        self.filtro_nivel = QComboBox()
        self.filtro_nivel.addItems(["Todos los niveles", "junior", "senior", "gerente", "socio"])
        self.filtro_nivel.setMinimumHeight(32)
        self.filtro_disp = QComboBox()
        self.filtro_disp.addItems(["Toda disponibilidad", "tiempo completo", "medio tiempo"])
        self.filtro_disp.setMinimumHeight(32)
        self.btn_buscar  = QPushButton("  Buscar")
        self.btn_limpiar = QPushButton("  Limpiar")
        self.btn_buscar.setMinimumHeight(32)
        self.btn_limpiar.setMinimumHeight(32)
        filtros.addWidget(self.campo_busqueda, 3)
        filtros.addWidget(self.filtro_nivel, 1)
        filtros.addWidget(self.filtro_disp, 1)
        filtros.addWidget(self.btn_buscar)
        filtros.addWidget(self.btn_limpiar)
        layout.addLayout(filtros)

        self.tabla = QTableWidget()
        columnas = ["Codigo", "Nombres", "Apellidos", "Nivel",
                    "Especialidades", "Tarifa/Hora", "Idiomas", "Disponibilidad"]
        self.tabla.setColumnCount(len(columnas))
        self.tabla.setHorizontalHeaderLabels(columnas)
        self.tabla.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.tabla.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.tabla.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.tabla.horizontalHeader().setStretchLastSection(True)
        self.tabla.verticalHeader().setVisible(False)
        self.tabla.setAlternatingRowColors(True)
        layout.addWidget(self.tabla)

        self.lbl_total = QLabel("Total: 0 registros")
        layout.addWidget(self.lbl_total)

        self.btn_nuevo.clicked.connect(self._abrir_formulario_nuevo)
        self.btn_editar.clicked.connect(self._abrir_formulario_editar)
        self.btn_eliminar.clicked.connect(self._eliminar)
        self.btn_excel.clicked.connect(self._exportar_excel)
        self.btn_pdf.clicked.connect(self._exportar_pdf)
        self.btn_buscar.clicked.connect(self._buscar)
        self.btn_limpiar.clicked.connect(self._limpiar)
        self.campo_busqueda.returnPressed.connect(self._buscar)
        self.filtro_nivel.currentIndexChanged.connect(self._buscar)
        self.filtro_disp.currentIndexChanged.connect(self._buscar)

        iconos_crud(self.btn_nuevo, self.btn_editar, self.btn_eliminar,
                    self.btn_excel, self.btn_pdf, self.btn_buscar, self.btn_limpiar)

    def _aplicar_tema(self):
        t = obtener_tema()
        self.setStyleSheet(f"background-color: {t['fondo']}; color: {t['texto']};")
        self.titulo.setStyleSheet(f"font-size: 22px; font-weight: bold; color: {t['titulo']};")
        self.lbl_total.setStyleSheet(f"font-size: 11px; color: {t['texto_suave']};")
        self.btn_nuevo.setStyleSheet(estilo_boton(t["btn_primario"],  t["btn_primario_txt"]))
        self.btn_editar.setStyleSheet(estilo_boton(t["btn_editar"],   t["btn_editar_txt"]))
        self.btn_eliminar.setStyleSheet(estilo_boton(t["btn_peligro"],t["btn_peligro_txt"]))
        self.btn_excel.setStyleSheet(estilo_boton(t["btn_excel"],     t["btn_excel_txt"]))
        self.btn_pdf.setStyleSheet(estilo_boton(t["btn_pdf"],         t["btn_pdf_txt"]))
        self.btn_buscar.setStyleSheet(estilo_boton(t["btn_buscar"],   t["btn_buscar_txt"]))
        self.btn_limpiar.setStyleSheet(estilo_boton(t["btn_limpiar"], t["btn_limpiar_txt"]))
        self.campo_busqueda.setStyleSheet(estilo_input(t))
        self.filtro_nivel.setStyleSheet(estilo_input(t))
        self.filtro_disp.setStyleSheet(estilo_input(t))
        self.tabla.setStyleSheet(estilo_tabla(t))

    def _cargar_datos(self, consultores=None):
        if consultores is None:
            consultores = self.repo.select_all()
        self.tabla.setRowCount(len(consultores))
        for fila, c in enumerate(consultores):
            self.tabla.setItem(fila, 0, QTableWidgetItem(str(c.codigo_empleado or "")))
            self.tabla.setItem(fila, 1, QTableWidgetItem(str(c.nombres or "")))
            self.tabla.setItem(fila, 2, QTableWidgetItem(str(c.apellidos or "")))
            self.tabla.setItem(fila, 3, QTableWidgetItem(str(c.nivel or "")))
            self.tabla.setItem(fila, 4, QTableWidgetItem(str(c.especialidades or "")))
            self.tabla.setItem(fila, 5, QTableWidgetItem(f"${c.tarifa_horaria or 0:.2f}"))
            self.tabla.setItem(fila, 6, QTableWidgetItem(str(c.idiomas or "")))
            self.tabla.setItem(fila, 7, QTableWidgetItem(str(c.disponibilidad or "")))
        self.tabla.resizeColumnsToContents()
        self.lbl_total.setText(f"Total: {len(consultores)} registros")

    def _fila_seleccionada(self):
        fila = self.tabla.currentRow()
        if fila == -1:
            return None
        return self.tabla.item(fila, 0).text()

    def _buscar(self):
        termino    = self.campo_busqueda.text().strip()
        nivel      = self.filtro_nivel.currentText()
        disp       = self.filtro_disp.currentText()
        resultados = self.repo.search(termino) if termino else self.repo.select_all()
        if nivel != "Todos los niveles":
            resultados = [c for c in resultados if c.nivel == nivel]
        if disp  != "Toda disponibilidad":
            resultados = [c for c in resultados if c.disponibilidad == disp]
        self._cargar_datos(resultados)

    def _limpiar(self):
        self.campo_busqueda.clear()
        self.filtro_nivel.setCurrentIndex(0)
        self.filtro_disp.setCurrentIndex(0)
        self._cargar_datos()

    def _abrir_formulario_nuevo(self):
        dialogo = FormularioConsultor(self.repo)
        if dialogo.exec() == QDialog.DialogCode.Accepted:
            self._cargar_datos()

    def _abrir_formulario_editar(self):
        codigo = self._fila_seleccionada()
        if not codigo:
            QMessageBox.warning(self, "Aviso", "Selecciona un consultor para editar.")
            return
        consultor = self.repo.select_WHERE_codigo_empleado(codigo)
        dialogo   = FormularioConsultor(self.repo, consultor)
        if dialogo.exec() == QDialog.DialogCode.Accepted:
            self._cargar_datos()

    def _eliminar(self):
        codigo = self._fila_seleccionada()
        if not codigo:
            QMessageBox.warning(self, "Aviso", "Selecciona un consultor para eliminar.")
            return
        respuesta = QMessageBox.question(
            self, "Confirmar eliminacion",
            f"Estas seguro de eliminar el consultor {codigo}?\n\nEsta accion no se puede deshacer.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if respuesta == QMessageBox.StandardButton.Yes:
            try:
                self.repo.delete_consultor(codigo)
                for ext in [".jpg", ".png", ".gif"]:
                    ruta = os.path.join(CARPETA_IMAGENES, f"{codigo}{ext}")
                    if os.path.exists(ruta):
                        os.remove(ruta)
                self._cargar_datos()
                QMessageBox.information(self, "Exito", "Consultor eliminado correctamente.")
            except Exception as e:
                QMessageBox.critical(self, "No se puede eliminar", str(e))

    def _exportar_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            QMessageBox.critical(self, "Error", "Instala openpyxl:\npip install openpyxl")
            return

        dialogo = DialogoFiltroConsultores(self)
        if dialogo.exec() != QDialog.DialogCode.Accepted:
            return

        datos = self._aplicar_filtros(dialogo.obtener_filtros())
        ruta, _ = QFileDialog.getSaveFileName(self, "Guardar Excel", "consultores.xlsx", "Excel (*.xlsx)")
        if not ruta:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Consultores"
        fill_enc = PatternFill("solid", fgColor="14532D")
        font_enc = Font(bold=True, color="FFFFFF", size=11)
        font_tit = Font(bold=True, size=14, color="14532D")
        borde    = Border(left=Side(style="thin"), right=Side(style="thin"),
                          top=Side(style="thin"),  bottom=Side(style="thin"))
        fill_par = PatternFill("solid", fgColor="DCFCE7")

        ws.merge_cells("A1:H1")
        ws["A1"] = "Reporte de Consultores"
        ws["A1"].font      = font_tit
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:H2")
        ws["A2"] = f"Generado: {date.today().strftime('%d/%m/%Y')}  |  Total: {len(datos)} registros"
        ws["A2"].alignment = Alignment(horizontal="center")
        ws["A2"].font      = Font(italic=True, color="666666")

        cols = ["Codigo", "Nombres", "Apellidos", "Nivel",
                "Especialidades", "Tarifa/Hora", "Idiomas", "Disponibilidad"]
        for i, col in enumerate(cols, 1):
            cell           = ws.cell(row=4, column=i, value=col)
            cell.font      = font_enc
            cell.fill      = fill_enc
            cell.alignment = Alignment(horizontal="center")
            cell.border    = borde

        for idx, c in enumerate(datos, 5):
            fila_data = [c.codigo_empleado, c.nombres, c.apellidos, c.nivel,
                         c.especialidades, float(c.tarifa_horaria or 0),
                         c.idiomas, c.disponibilidad]
            fill = fill_par if idx % 2 == 0 else None
            for j, val in enumerate(fila_data, 1):
                cell        = ws.cell(row=idx, column=j, value=val or "")
                cell.border = borde
                if fill:
                    cell.fill = fill

        anchos = [12, 20, 20, 10, 30, 12, 20, 16]
        for i, ancho in enumerate(anchos, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

        wb.save(ruta)
        QMessageBox.information(self, "Exportado", f"Excel guardado en:\n{ruta}")

    def _exportar_pdf(self):
        try:
            from reportlab.lib.pagesizes import landscape, A4
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import ParagraphStyle
            from reportlab.lib.units import cm
        except ImportError:
            QMessageBox.critical(self, "Error", "Instala reportlab:\npip install reportlab")
            return

        dialogo = DialogoFiltroConsultores(self)
        if dialogo.exec() != QDialog.DialogCode.Accepted:
            return

        datos   = self._aplicar_filtros(dialogo.obtener_filtros())
        ruta, _ = QFileDialog.getSaveFileName(self, "Guardar PDF", "consultores.pdf", "PDF (*.pdf)")
        if not ruta:
            return

        doc   = SimpleDocTemplate(ruta, pagesize=landscape(A4),
                                  topMargin=1.5*cm, bottomMargin=1.5*cm)
        story = []
        st = ParagraphStyle("t", fontSize=16, fontName="Helvetica-Bold",
                             spaceAfter=4, textColor=colors.HexColor("#14532D"))
        ss = ParagraphStyle("s", fontSize=9, fontName="Helvetica",
                             spaceAfter=12, textColor=colors.grey)
        story.append(Paragraph("Reporte de Consultores", st))
        story.append(Paragraph(
            f"Generado: {date.today().strftime('%d/%m/%Y')}  |  Total: {len(datos)} registros", ss))
        story.append(Spacer(1, 0.3*cm))

        enc = ["Codigo", "Nombres", "Apellidos", "Nivel",
               "Especialidades", "Tarifa/Hora", "Disponibilidad"]
        filas = [enc]
        for c in datos:
            filas.append([
                c.codigo_empleado or "", c.nombres or "", c.apellidos or "",
                c.nivel or "", c.especialidades or "",
                f"${c.tarifa_horaria or 0:.2f}", c.disponibilidad or ""
            ])

        tabla = Table(filas, repeatRows=1)
        tabla.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),(-1,0),  colors.HexColor("#14532D")),
            ("TEXTCOLOR",     (0,0),(-1,0),  colors.white),
            ("FONTNAME",      (0,0),(-1,0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0,0),(-1,-1), 8),
            ("ALIGN",         (0,0),(-1,-1), "CENTER"),
            ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, colors.HexColor("#DCFCE7")]),
            ("GRID",          (0,0),(-1,-1), 0.5, colors.HexColor("#CCCCCC")),
            ("TOPPADDING",    (0,0),(-1,-1), 5),
            ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ]))
        story.append(tabla)
        doc.build(story)
        QMessageBox.information(self, "Exportado", f"PDF guardado en:\n{ruta}")

    def _aplicar_filtros(self, filtros):
        nivel = filtros.get("nivel")
        disp  = filtros.get("disponibilidad")
        datos = self.repo.select_all()
        if nivel:
            datos = [c for c in datos if c.nivel == nivel]
        if disp:
            datos = [c for c in datos if c.disponibilidad == disp]
        return datos


class DialogoFiltroConsultores(QDialog):

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Filtros de exportacion")
        self.setWindowIcon(favicon("consultores"))
        self.setMinimumWidth(320)
        t = obtener_tema()
        self.setStyleSheet(f"background-color: {t['fondo_panel']}; color: {t['texto']};")
        layout = QVBoxLayout()
        self.setLayout(layout)

        layout.addWidget(QLabel("Filtrar por nivel:"))
        self.combo_nivel = QComboBox()
        self.combo_nivel.addItems(["Todos", "junior", "senior", "gerente", "socio"])
        self.combo_nivel.setStyleSheet(estilo_input(t))
        layout.addWidget(self.combo_nivel)

        layout.addWidget(QLabel("Filtrar por disponibilidad:"))
        self.combo_disp = QComboBox()
        self.combo_disp.addItems(["Todos", "tiempo completo", "medio tiempo"])
        self.combo_disp.setStyleSheet(estilo_input(t))
        layout.addWidget(self.combo_disp)

        barra      = QHBoxLayout()
        btn_ok     = QPushButton("  Exportar")
        btn_cancel = QPushButton("  Cancelar")
        btn_ok.setStyleSheet(estilo_boton(t["btn_primario"], t["btn_primario_txt"]))
        btn_cancel.setStyleSheet(estilo_boton(t["btn_limpiar"], t["btn_limpiar_txt"]))
        barra.addStretch()
        barra.addWidget(btn_ok)
        barra.addWidget(btn_cancel)
        layout.addLayout(barra)
        btn_ok.clicked.connect(self.accept)
        btn_cancel.clicked.connect(self.reject)

    def obtener_filtros(self):
        nivel = self.combo_nivel.currentText()
        disp  = self.combo_disp.currentText()
        return {
            "nivel":         None if nivel == "Todos" else nivel,
            "disponibilidad":None if disp  == "Todos" else disp,
        }


class FormularioConsultor(QDialog):

    def __init__(self, repo, consultor=None):
        super().__init__()
        self.repo        = repo
        self.consultor   = consultor
        self.editando    = consultor is not None
        self.ruta_imagen = None
        t                = obtener_tema()

        self.setWindowTitle("Editar Consultor" if self.editando else "Nuevo Consultor")
        self.setWindowIcon(favicon("consultores"))
        self.setMinimumWidth(560)
        self.setStyleSheet(f"background-color: {t['fondo_panel']}; color: {t['texto']};")

        layout_principal = QHBoxLayout()
        self.setLayout(layout_principal)

        scroll     = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("border: none;")
        contenedor = QWidget()
        scroll.setWidget(contenedor)
        form_layout = QVBoxLayout()
        contenedor.setLayout(form_layout)

        form   = QFormLayout()
        form.setSpacing(10)
        estilo = estilo_input(t)

        self.campo_codigo   = QLineEdit()
        self.campo_nombres  = QLineEdit()
        self.campo_apellidos= QLineEdit()
        self.campo_doc      = QLineEdit()
        self.campo_formacion= QLineEdit()
        self.campo_certif   = QLineEdit()
        self.campo_espec    = QLineEdit()
        self.campo_anios    = QSpinBox()
        self.campo_anios.setRange(0, 50)
        self.campo_nivel    = QComboBox()
        self.campo_nivel.addItems(["junior", "senior", "gerente", "socio"])
        self.campo_tarifa   = QDoubleSpinBox()
        self.campo_tarifa.setRange(0, 9999.99)
        self.campo_tarifa.setDecimals(2)
        self.campo_tarifa.setPrefix("$ ")
        self.campo_idiomas  = QLineEdit()
        self.campo_disp     = QComboBox()
        self.campo_disp.addItems(["tiempo completo", "medio tiempo", "no disponible"])

        for w in [self.campo_codigo, self.campo_nombres, self.campo_apellidos,
                  self.campo_doc, self.campo_formacion, self.campo_certif,
                  self.campo_espec, self.campo_anios, self.campo_nivel,
                  self.campo_tarifa, self.campo_idiomas, self.campo_disp]:
            w.setStyleSheet(estilo)

        form.addRow("Codigo *:",         self.campo_codigo)
        form.addRow("Nombres *:",        self.campo_nombres)
        form.addRow("Apellidos *:",      self.campo_apellidos)
        form.addRow("Documento:",        self.campo_doc)
        form.addRow("Formacion:",        self.campo_formacion)
        form.addRow("Certificaciones:",  self.campo_certif)
        form.addRow("Especialidades:",   self.campo_espec)
        form.addRow("Anos experiencia:", self.campo_anios)
        form.addRow("Nivel *:",          self.campo_nivel)
        form.addRow("Tarifa/Hora:",      self.campo_tarifa)
        form.addRow("Idiomas:",          self.campo_idiomas)
        form.addRow("Disponibilidad *:", self.campo_disp)
        form_layout.addLayout(form)

        barra = QHBoxLayout()
        self.btn_guardar  = QPushButton("  Guardar")
        self.btn_cancelar = QPushButton("  Cancelar")
        self.btn_guardar.setMinimumHeight(34)
        self.btn_cancelar.setMinimumHeight(34)
        self.btn_guardar.setStyleSheet(estilo_boton(t["btn_primario"],  t["btn_primario_txt"]))
        self.btn_cancelar.setStyleSheet(estilo_boton(t["btn_limpiar"],  t["btn_limpiar_txt"]))
        barra.addStretch()
        barra.addWidget(self.btn_guardar)
        barra.addWidget(self.btn_cancelar)
        form_layout.addLayout(barra)

        self.btn_guardar.clicked.connect(self._guardar)
        self.btn_cancelar.clicked.connect(self.reject)
        layout_principal.addWidget(scroll, 2)

        panel_imagen = QFrame()
        panel_imagen.setFixedWidth(180)
        panel_imagen.setStyleSheet(
            f"background-color: {t['fondo_input']}; "
            f"border-radius: 8px; border: 1px solid {t['borde']};"
        )
        img_layout = QVBoxLayout()
        img_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        panel_imagen.setLayout(img_layout)

        lbl_img_titulo = QLabel("Foto del Consultor")
        lbl_img_titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_img_titulo.setStyleSheet(
            f"color: {t['texto_suave']}; font-size: 11px; font-weight: bold; border: none;"
        )
        img_layout.addWidget(lbl_img_titulo)

        self.lbl_imagen = QLabel()
        self.lbl_imagen.setFixedSize(150, 150)
        self.lbl_imagen.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_imagen.setStyleSheet(
            f"border: 2px dashed {t['borde']}; border-radius: 75px; color: {t['texto_suave']};"
        )
        self.lbl_imagen.setText("Sin foto")
        img_layout.addWidget(self.lbl_imagen)

        self.btn_cargar       = QPushButton("  Cargar foto")
        self.btn_eliminar_img = QPushButton("  Quitar foto")
        self.btn_cargar.setStyleSheet(estilo_boton(t["btn_buscar"],  t["btn_buscar_txt"]))
        self.btn_eliminar_img.setStyleSheet(estilo_boton(t["btn_peligro"], t["btn_peligro_txt"]))
        img_layout.addWidget(self.btn_cargar)
        img_layout.addWidget(self.btn_eliminar_img)

        lbl_formatos = QLabel("JPG, PNG, GIF\nMax: 5MB")
        lbl_formatos.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_formatos.setStyleSheet(f"color: {t['texto_suave']}; font-size: 10px; border: none;")
        img_layout.addWidget(lbl_formatos)
        img_layout.addStretch()

        self.btn_cargar.clicked.connect(self._cargar_imagen)
        self.btn_eliminar_img.clicked.connect(self._quitar_imagen)
        layout_principal.addWidget(panel_imagen, 1)

        iconos_formulario(self.btn_guardar, self.btn_cancelar, self.btn_cargar)

        if self.editando:
            self.campo_codigo.setText(consultor.codigo_empleado)
            self.campo_codigo.setEnabled(False)
            self.campo_nombres.setText(consultor.nombres or "")
            self.campo_apellidos.setText(consultor.apellidos or "")
            self.campo_doc.setText(consultor.documento_identidad or "")
            self.campo_formacion.setText(consultor.formacion_academica or "")
            self.campo_certif.setText(consultor.certificaciones or "")
            self.campo_espec.setText(consultor.especialidades or "")
            self.campo_anios.setValue(consultor.anios_experiencia or 0)
            self.campo_nivel.setCurrentText(consultor.nivel or "junior")
            self.campo_tarifa.setValue(float(consultor.tarifa_horaria or 0))
            self.campo_idiomas.setText(consultor.idiomas or "")
            self.campo_disp.setCurrentText(consultor.disponibilidad or "tiempo completo")
            self._cargar_imagen_existente(consultor.codigo_empleado)

    def _cargar_imagen_existente(self, codigo):
        for ext in [".jpg", ".png", ".gif"]:
            ruta = os.path.join(CARPETA_IMAGENES, f"{codigo}{ext}")
            if os.path.exists(ruta):
                self._mostrar_imagen(ruta)
                self.ruta_imagen = ruta
                break

    def _cargar_imagen(self):
        ruta, _ = QFileDialog.getOpenFileName(
            self, "Seleccionar foto", "", "Imagenes (*.jpg *.jpeg *.png *.gif)"
        )
        if not ruta:
            return
        if os.path.getsize(ruta) > 5 * 1024 * 1024:
            QMessageBox.warning(self, "Error", "La imagen supera el tamaño maximo de 5MB.")
            return
        try:
            from PIL import Image
            img = Image.open(ruta)
            if img.format not in ["JPEG", "PNG", "GIF"]:
                QMessageBox.warning(self, "Error", "Formato no soportado. Usa JPG, PNG o GIF.")
                return
        except Exception:
            QMessageBox.warning(self, "Error", "No se pudo leer la imagen.")
            return
        self.ruta_imagen = ruta
        self._mostrar_imagen(ruta)

    def _mostrar_imagen(self, ruta):
        pixmap = QPixmap(ruta).scaled(
            150, 150,
            Qt.AspectRatioMode.KeepAspectRatio,
            Qt.TransformationMode.SmoothTransformation
        )
        self.lbl_imagen.setPixmap(pixmap)
        self.lbl_imagen.setText("")

    def _quitar_imagen(self):
        self.ruta_imagen = None
        self.lbl_imagen.setPixmap(QPixmap())
        self.lbl_imagen.setText("Sin foto")

    def _validar(self):
        if not self.campo_codigo.text().strip():
            QMessageBox.warning(self, "Error", "El codigo es obligatorio.")
            return False
        if len(self.campo_codigo.text().strip()) < 3:
            QMessageBox.warning(self, "Error", "El codigo debe tener al menos 3 caracteres.")
            return False
        if not self.campo_nombres.text().strip():
            QMessageBox.warning(self, "Error", "El nombre es obligatorio.")
            return False
        if not self.campo_apellidos.text().strip():
            QMessageBox.warning(self, "Error", "El apellido es obligatorio.")
            return False
        if self.campo_tarifa.value() < 0:
            QMessageBox.warning(self, "Error", "La tarifa no puede ser negativa.")
            return False
        return True

    def _guardar(self):
        if not self._validar():
            return
        accion = "actualizar" if self.editando else "guardar"
        respuesta = QMessageBox.question(
            self, "Confirmar",
            f"Deseas {accion} este consultor?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if respuesta != QMessageBox.StandardButton.Yes:
            return

        consultor = Consultor(
            codigo_empleado     = self.campo_codigo.text().strip(),
            nombres             = self.campo_nombres.text().strip(),
            apellidos           = self.campo_apellidos.text().strip(),
            documento_identidad = self.campo_doc.text().strip() or None,
            formacion_academica = self.campo_formacion.text().strip() or None,
            certificaciones     = self.campo_certif.text().strip() or None,
            especialidades      = self.campo_espec.text().strip() or None,
            anios_experiencia   = self.campo_anios.value(),
            nivel               = self.campo_nivel.currentText(),
            tarifa_horaria      = self.campo_tarifa.value(),
            idiomas             = self.campo_idiomas.text().strip() or None,
            disponibilidad      = self.campo_disp.currentText()
        )

        try:
            if self.editando:
                self.repo.update(consultor)
            else:
                self.repo.insert(consultor)
            if self.ruta_imagen:
                ext     = os.path.splitext(self.ruta_imagen)[1].lower()
                destino = os.path.join(CARPETA_IMAGENES, f"{consultor.codigo_empleado}{ext}")
                if self.ruta_imagen != destino:
                    shutil.copy2(self.ruta_imagen, destino)
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Error al guardar", str(e))

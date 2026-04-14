from datetime import date

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QTableWidget,
    QTableWidgetItem, QDialog, QFormLayout, QLineEdit, QComboBox,
    QLabel, QMessageBox, QHeaderView, QFileDialog, QScrollArea,
    QTextEdit, QDoubleSpinBox
)
from PyQt6.QtCore import Qt

from models.servicio import Servicio
from ui.themes import obtener_tema, estilo_boton, estilo_input, estilo_tabla
from ui.iconos.iconos import iconos_crud, iconos_formulario, favicon


class ServiciosWidget(QWidget):

    def __init__(self, repo):
        super().__init__()
        self.repo = repo
        self._construir_ui()
        self._aplicar_tema()
        self._cargar_datos()

    def _construir_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(16, 16, 16, 16)
        self.setLayout(layout)

        self.titulo = QLabel("Servicios")
        self.titulo.setStyleSheet("font-size: 22px; font-weight: bold; padding: 4px 0;")
        layout.addWidget(self.titulo)

        barra = QHBoxLayout()
        self.btn_nuevo    = QPushButton("  Nuevo")
        self.btn_editar   = QPushButton("  Editar")
        self.btn_eliminar = QPushButton("  Eliminar")
        self.btn_excel    = QPushButton("  Exportar Excel")
        self.btn_pdf      = QPushButton("  Exportar PDF")
        for btn in [self.btn_nuevo, self.btn_editar, self.btn_eliminar,
                    self.btn_excel, self.btn_pdf]:
            btn.setMinimumHeight(34)
        barra.addWidget(self.btn_nuevo)
        barra.addWidget(self.btn_editar)
        barra.addWidget(self.btn_eliminar)
        barra.addSpacing(20)
        barra.addWidget(self.btn_excel)
        barra.addWidget(self.btn_pdf)
        barra.addStretch()
        layout.addLayout(barra)

        filtros = QHBoxLayout()
        self.campo_busqueda = QLineEdit()
        self.campo_busqueda.setPlaceholderText("Buscar por nombre o categoria...")
        self.campo_busqueda.setMinimumHeight(32)
        self.filtro_categoria = QComboBox()
        self.filtro_categoria.addItems([
            "Todas las categorias", "Estrategia", "Finanzas",
            "Operaciones", "Marketing", "Recursos Humanos", "Tecnologia"
        ])
        self.filtro_categoria.setMinimumHeight(32)
        self.btn_buscar  = QPushButton("  Buscar")
        self.btn_limpiar = QPushButton("  Limpiar")
        self.btn_buscar.setMinimumHeight(32)
        self.btn_limpiar.setMinimumHeight(32)
        filtros.addWidget(self.campo_busqueda, 3)
        filtros.addWidget(self.filtro_categoria, 1)
        filtros.addWidget(self.btn_buscar)
        filtros.addWidget(self.btn_limpiar)
        layout.addLayout(filtros)

        self.tabla = QTableWidget()
        columnas = ["Codigo", "Nombre Comercial", "Categoria",
                    "Duracion Estimada", "Metodologia", "Tarifario"]
        self.tabla.setColumnCount(len(columnas))
        self.tabla.setHorizontalHeaderLabels(columnas)
        self.tabla.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.tabla.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.tabla.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.tabla.horizontalHeader().setStretchLastSection(True)
        self.tabla.verticalHeader().setVisible(False)
        self.tabla.setAlternatingRowColors(True)
        layout.addWidget(self.tabla)

        self.lbl_total = QLabel("Total: 0 registros")
        layout.addWidget(self.lbl_total)

        self.btn_nuevo.clicked.connect(self._abrir_formulario_nuevo)
        self.btn_editar.clicked.connect(self._abrir_formulario_editar)
        self.btn_eliminar.clicked.connect(self._eliminar)
        self.btn_excel.clicked.connect(self._exportar_excel)
        self.btn_pdf.clicked.connect(self._exportar_pdf)
        self.btn_buscar.clicked.connect(self._buscar)
        self.btn_limpiar.clicked.connect(self._limpiar)
        self.campo_busqueda.returnPressed.connect(self._buscar)
        self.filtro_categoria.currentIndexChanged.connect(self._buscar)

        iconos_crud(self.btn_nuevo, self.btn_editar, self.btn_eliminar,
                    self.btn_excel, self.btn_pdf, self.btn_buscar, self.btn_limpiar)

    def _aplicar_tema(self):
        t = obtener_tema()
        self.setStyleSheet(f"background-color: {t['fondo']}; color: {t['texto']};")
        self.titulo.setStyleSheet(f"font-size: 22px; font-weight: bold; color: {t['titulo']};")
        self.lbl_total.setStyleSheet(f"font-size: 11px; color: {t['texto_suave']};")
        self.btn_nuevo.setStyleSheet(estilo_boton(t["btn_primario"],  t["btn_primario_txt"]))
        self.btn_editar.setStyleSheet(estilo_boton(t["btn_editar"],   t["btn_editar_txt"]))
        self.btn_eliminar.setStyleSheet(estilo_boton(t["btn_peligro"],t["btn_peligro_txt"]))
        self.btn_excel.setStyleSheet(estilo_boton(t["btn_excel"],     t["btn_excel_txt"]))
        self.btn_pdf.setStyleSheet(estilo_boton(t["btn_pdf"],         t["btn_pdf_txt"]))
        self.btn_buscar.setStyleSheet(estilo_boton(t["btn_buscar"],   t["btn_buscar_txt"]))
        self.btn_limpiar.setStyleSheet(estilo_boton(t["btn_limpiar"], t["btn_limpiar_txt"]))
        self.campo_busqueda.setStyleSheet(estilo_input(t))
        self.filtro_categoria.setStyleSheet(estilo_input(t))
        self.tabla.setStyleSheet(estilo_tabla(t))

    def _cargar_datos(self, servicios=None):
        if servicios is None:
            servicios = self.repo.select_all()
        self.tabla.setRowCount(len(servicios))
        for fila, s in enumerate(servicios):
            self.tabla.setItem(fila, 0, QTableWidgetItem(str(s.codigo or "")))
            self.tabla.setItem(fila, 1, QTableWidgetItem(str(s.nombre_comercial or "")))
            self.tabla.setItem(fila, 2, QTableWidgetItem(str(s.categoria or "")))
            self.tabla.setItem(fila, 3, QTableWidgetItem(str(s.duracion_estimada or "")))
            self.tabla.setItem(fila, 4, QTableWidgetItem(str(s.metodologia or "")))
            self.tabla.setItem(fila, 5, QTableWidgetItem(f"${s.tarifario_referencial or 0:,.2f}"))
        self.tabla.resizeColumnsToContents()
        self.lbl_total.setText(f"Total: {len(servicios)} registros")

    def _fila_seleccionada(self):
        fila = self.tabla.currentRow()
        if fila == -1:
            return None
        return self.tabla.item(fila, 0).text()

    def _buscar(self):
        termino    = self.campo_busqueda.text().strip()
        categoria  = self.filtro_categoria.currentText()
        resultados = self.repo.search(termino) if termino else self.repo.select_all()
        if categoria != "Todas las categorias":
            resultados = [s for s in resultados if s.categoria == categoria]
        self._cargar_datos(resultados)

    def _limpiar(self):
        self.campo_busqueda.clear()
        self.filtro_categoria.setCurrentIndex(0)
        self._cargar_datos()

    def _abrir_formulario_nuevo(self):
        dialogo = FormularioServicio(self.repo)
        if dialogo.exec() == QDialog.DialogCode.Accepted:
            self._cargar_datos()

    def _abrir_formulario_editar(self):
        codigo = self._fila_seleccionada()
        if not codigo:
            QMessageBox.warning(self, "Aviso", "Selecciona un servicio para editar.")
            return
        servicio = self.repo.select_WHERE_codigo(codigo)
        dialogo  = FormularioServicio(self.repo, servicio)
        if dialogo.exec() == QDialog.DialogCode.Accepted:
            self._cargar_datos()

    def _eliminar(self):
        codigo = self._fila_seleccionada()
        if not codigo:
            QMessageBox.warning(self, "Aviso", "Selecciona un servicio para eliminar.")
            return
        respuesta = QMessageBox.question(
            self, "Confirmar eliminacion",
            f"Estas seguro de eliminar el servicio {codigo}?\n\nEsta accion no se puede deshacer.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if respuesta == QMessageBox.StandardButton.Yes:
            try:
                self.repo.delete_servicio(codigo)
                self._cargar_datos()
                QMessageBox.information(self, "Exito", "Servicio eliminado correctamente.")
            except Exception as e:
                QMessageBox.critical(self, "No se puede eliminar", str(e))

    def _exportar_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            QMessageBox.critical(self, "Error", "Instala openpyxl:\npip install openpyxl")
            return

        dialogo = DialogoFiltroServicios(self)
        if dialogo.exec() != QDialog.DialogCode.Accepted:
            return

        datos   = self._aplicar_filtros(dialogo.obtener_filtros())
        ruta, _ = QFileDialog.getSaveFileName(self, "Guardar Excel", "servicios.xlsx", "Excel (*.xlsx)")
        if not ruta:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Servicios"
        fill_enc = PatternFill("solid", fgColor="78350F")
        font_enc = Font(bold=True, color="FFFFFF", size=11)
        font_tit = Font(bold=True, size=14, color="78350F")
        borde    = Border(left=Side(style="thin"), right=Side(style="thin"),
                          top=Side(style="thin"),  bottom=Side(style="thin"))
        fill_par = PatternFill("solid", fgColor="FEF3C7")

        ws.merge_cells("A1:G1")
        ws["A1"] = "Reporte de Servicios"
        ws["A1"].font      = font_tit
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:G2")
        ws["A2"] = f"Generado: {date.today().strftime('%d/%m/%Y')}  |  Total: {len(datos)} registros"
        ws["A2"].alignment = Alignment(horizontal="center")
        ws["A2"].font      = Font(italic=True, color="666666")

        cols = ["Codigo", "Nombre Comercial", "Categoria", "Duracion",
                "Metodologia", "Equipo Minimo", "Tarifario"]
        for i, col in enumerate(cols, 1):
            cell           = ws.cell(row=4, column=i, value=col)
            cell.font      = font_enc
            cell.fill      = fill_enc
            cell.alignment = Alignment(horizontal="center")
            cell.border    = borde

        for idx, s in enumerate(datos, 5):
            fila_data = [s.codigo, s.nombre_comercial, s.categoria,
                         s.duracion_estimada, s.metodologia, s.equipo_minimo,
                         float(s.tarifario_referencial or 0)]
            fill = fill_par if idx % 2 == 0 else None
            for j, val in enumerate(fila_data, 1):
                cell        = ws.cell(row=idx, column=j, value=val or "")
                cell.border = borde
                if fill:
                    cell.fill = fill

        anchos = [12, 30, 16, 16, 25, 20, 14]
        for i, ancho in enumerate(anchos, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

        wb.save(ruta)
        QMessageBox.information(self, "Exportado", f"Excel guardado en:\n{ruta}")

    def _exportar_pdf(self):
        try:
            from reportlab.lib.pagesizes import landscape, A4
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import ParagraphStyle
            from reportlab.lib.units import cm
        except ImportError:
            QMessageBox.critical(self, "Error", "Instala reportlab:\npip install reportlab")
            return

        dialogo = DialogoFiltroServicios(self)
        if dialogo.exec() != QDialog.DialogCode.Accepted:
            return

        datos   = self._aplicar_filtros(dialogo.obtener_filtros())
        ruta, _ = QFileDialog.getSaveFileName(self, "Guardar PDF", "servicios.pdf", "PDF (*.pdf)")
        if not ruta:
            return

        doc   = SimpleDocTemplate(ruta, pagesize=landscape(A4),
                                  topMargin=1.5*cm, bottomMargin=1.5*cm)
        story = []
        st = ParagraphStyle("t", fontSize=16, fontName="Helvetica-Bold",
                             spaceAfter=4, textColor=colors.HexColor("#78350F"))
        ss = ParagraphStyle("s", fontSize=9, fontName="Helvetica",
                             spaceAfter=12, textColor=colors.grey)
        story.append(Paragraph("Reporte de Servicios", st))
        story.append(Paragraph(
            f"Generado: {date.today().strftime('%d/%m/%Y')}  |  Total: {len(datos)} registros", ss))
        story.append(Spacer(1, 0.3*cm))

        enc   = ["Codigo", "Nombre Comercial", "Categoria",
                 "Duracion", "Metodologia", "Tarifario"]
        filas = [enc]
        for s in datos:
            filas.append([
                s.codigo or "", s.nombre_comercial or "", s.categoria or "",
                s.duracion_estimada or "", s.metodologia or "",
                f"${s.tarifario_referencial or 0:,.2f}"
            ])

        tabla = Table(filas, repeatRows=1)
        tabla.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),(-1,0),  colors.HexColor("#78350F")),
            ("TEXTCOLOR",     (0,0),(-1,0),  colors.white),
            ("FONTNAME",      (0,0),(-1,0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0,0),(-1,-1), 8),
            ("ALIGN",         (0,0),(-1,-1), "CENTER"),
            ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, colors.HexColor("#FEF3C7")]),
            ("GRID",          (0,0),(-1,-1), 0.5, colors.HexColor("#CCCCCC")),
            ("TOPPADDING",    (0,0),(-1,-1), 5),
            ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ]))
        story.append(tabla)
        doc.build(story)
        QMessageBox.information(self, "Exportado", f"PDF guardado en:\n{ruta}")

    def _aplicar_filtros(self, filtros):
        categoria = filtros.get("categoria")
        datos     = self.repo.select_all()
        if categoria:
            datos = [s for s in datos if s.categoria == categoria]
        return datos


class DialogoFiltroServicios(QDialog):

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Filtros de exportacion")
        self.setWindowIcon(favicon("servicios"))
        self.setMinimumWidth(300)
        t = obtener_tema()
        self.setStyleSheet(f"background-color: {t['fondo_panel']}; color: {t['texto']};")
        layout = QVBoxLayout()
        self.setLayout(layout)

        layout.addWidget(QLabel("Filtrar por categoria:"))
        self.combo_cat = QComboBox()
        self.combo_cat.addItems([
            "Todas", "Estrategia", "Finanzas",
            "Operaciones", "Marketing", "Recursos Humanos", "Tecnologia"
        ])
        self.combo_cat.setStyleSheet(estilo_input(t))
        layout.addWidget(self.combo_cat)

        barra      = QHBoxLayout()
        btn_ok     = QPushButton("  Exportar")
        btn_cancel = QPushButton("  Cancelar")
        btn_ok.setStyleSheet(estilo_boton(t["btn_primario"], t["btn_primario_txt"]))
        btn_cancel.setStyleSheet(estilo_boton(t["btn_limpiar"], t["btn_limpiar_txt"]))
        barra.addStretch()
        barra.addWidget(btn_ok)
        barra.addWidget(btn_cancel)
        layout.addLayout(barra)
        btn_ok.clicked.connect(self.accept)
        btn_cancel.clicked.connect(self.reject)

    def obtener_filtros(self):
        cat = self.combo_cat.currentText()
        return {"categoria": None if cat == "Todas" else cat}


class FormularioServicio(QDialog):

    def __init__(self, repo, servicio=None):
        super().__init__()
        self.repo     = repo
        self.servicio = servicio
        self.editando = servicio is not None
        t             = obtener_tema()

        self.setWindowTitle("Editar Servicio" if self.editando else "Nuevo Servicio")
        self.setWindowIcon(favicon("servicios"))
        self.setMinimumWidth(500)
        self.setStyleSheet(f"background-color: {t['fondo_panel']}; color: {t['texto']};")

        layout = QVBoxLayout()
        self.setLayout(layout)

        scroll     = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("border: none;")
        contenedor = QWidget()
        scroll.setWidget(contenedor)
        form_layout = QVBoxLayout()
        contenedor.setLayout(form_layout)

        form   = QFormLayout()
        form.setSpacing(10)
        estilo = estilo_input(t)

        self.campo_codigo    = QLineEdit()
        self.campo_nombre    = QLineEdit()
        self.campo_categoria = QComboBox()
        self.campo_categoria.addItems([
            "Estrategia", "Finanzas", "Operaciones",
            "Marketing", "Recursos Humanos", "Tecnologia"
        ])
        self.campo_descripcion  = QTextEdit()
        self.campo_descripcion.setMaximumHeight(80)
        self.campo_entregables  = QTextEdit()
        self.campo_entregables.setMaximumHeight(60)
        self.campo_duracion     = QLineEdit()
        self.campo_metodologia  = QLineEdit()
        self.campo_beneficios   = QTextEdit()
        self.campo_beneficios.setMaximumHeight(60)
        self.campo_equipo       = QLineEdit()
        self.campo_tarifario    = QDoubleSpinBox()
        self.campo_tarifario.setRange(0, 9999999.99)
        self.campo_tarifario.setDecimals(2)
        self.campo_tarifario.setPrefix("$ ")
        self.campo_casos        = QTextEdit()
        self.campo_casos.setMaximumHeight(60)

        for w in [self.campo_codigo, self.campo_nombre, self.campo_categoria,
                  self.campo_descripcion, self.campo_entregables, self.campo_duracion,
                  self.campo_metodologia, self.campo_beneficios, self.campo_equipo,
                  self.campo_tarifario, self.campo_casos]:
            w.setStyleSheet(estilo)

        form.addRow("Codigo *:",         self.campo_codigo)
        form.addRow("Nombre *:",         self.campo_nombre)
        form.addRow("Categoria *:",      self.campo_categoria)
        form.addRow("Descripcion:",      self.campo_descripcion)
        form.addRow("Entregables:",      self.campo_entregables)
        form.addRow("Duracion:",         self.campo_duracion)
        form.addRow("Metodologia:",      self.campo_metodologia)
        form.addRow("Beneficios:",       self.campo_beneficios)
        form.addRow("Equipo Minimo:",    self.campo_equipo)
        form.addRow("Tarifario:",        self.campo_tarifario)
        form.addRow("Casos de Exito:",   self.campo_casos)
        form_layout.addLayout(form)

        barra = QHBoxLayout()
        self.btn_guardar  = QPushButton("  Guardar")
        self.btn_cancelar = QPushButton("  Cancelar")
        self.btn_guardar.setMinimumHeight(34)
        self.btn_cancelar.setMinimumHeight(34)
        self.btn_guardar.setStyleSheet(estilo_boton(t["btn_primario"],  t["btn_primario_txt"]))
        self.btn_cancelar.setStyleSheet(estilo_boton(t["btn_limpiar"],  t["btn_limpiar_txt"]))
        barra.addStretch()
        barra.addWidget(self.btn_guardar)
        barra.addWidget(self.btn_cancelar)
        form_layout.addLayout(barra)

        self.btn_guardar.clicked.connect(self._guardar)
        self.btn_cancelar.clicked.connect(self.reject)
        layout.addWidget(scroll)

        iconos_formulario(self.btn_guardar, self.btn_cancelar)

        if self.editando:
            self.campo_codigo.setText(servicio.codigo)
            self.campo_codigo.setEnabled(False)
            self.campo_nombre.setText(servicio.nombre_comercial or "")
            self.campo_categoria.setCurrentText(servicio.categoria or "Estrategia")
            self.campo_descripcion.setPlainText(servicio.descripcion or "")
            self.campo_entregables.setPlainText(servicio.entregables_tipicos or "")
            self.campo_duracion.setText(servicio.duracion_estimada or "")
            self.campo_metodologia.setText(servicio.metodologia or "")
            self.campo_beneficios.setPlainText(servicio.beneficios_cliente or "")
            self.campo_equipo.setText(servicio.equipo_minimo or "")
            self.campo_tarifario.setValue(float(servicio.tarifario_referencial or 0))
            self.campo_casos.setPlainText(servicio.casos_exito or "")

    def _validar(self):
        if not self.campo_codigo.text().strip():
            QMessageBox.warning(self, "Error", "El codigo es obligatorio.")
            return False
        if len(self.campo_codigo.text().strip()) < 3:
            QMessageBox.warning(self, "Error", "El codigo debe tener al menos 3 caracteres.")
            return False
        if not self.campo_nombre.text().strip():
            QMessageBox.warning(self, "Error", "El nombre comercial es obligatorio.")
            return False
        if self.campo_tarifario.value() < 0:
            QMessageBox.warning(self, "Error", "El tarifario no puede ser negativo.")
            return False
        return True

    def _guardar(self):
        if not self._validar():
            return
        accion = "actualizar" if self.editando else "guardar"
        respuesta = QMessageBox.question(
            self, "Confirmar",
            f"Deseas {accion} este servicio?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if respuesta != QMessageBox.StandardButton.Yes:
            return

        servicio = Servicio(
            codigo                = self.campo_codigo.text().strip(),
            nombre_comercial      = self.campo_nombre.text().strip(),
            categoria             = self.campo_categoria.currentText(),
            descripcion           = self.campo_descripcion.toPlainText().strip() or None,
            entregables_tipicos   = self.campo_entregables.toPlainText().strip() or None,
            duracion_estimada     = self.campo_duracion.text().strip() or None,
            metodologia           = self.campo_metodologia.text().strip() or None,
            beneficios_cliente    = self.campo_beneficios.toPlainText().strip() or None,
            equipo_minimo         = self.campo_equipo.text().strip() or None,
            tarifario_referencial = self.campo_tarifario.value(),
            casos_exito           = self.campo_casos.toPlainText().strip() or None
        )

        try:
            if self.editando:
                self.repo.update(servicio)
            else:
                self.repo.insert(servicio)
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Error al guardar", str(e))

import os
import re
import shutil
from datetime import date

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QTableWidget,
    QTableWidgetItem, QDialog, QFormLayout, QLineEdit, QComboBox,
    QLabel, QMessageBox, QHeaderView, QDateEdit, QFileDialog,
    QScrollArea, QFrame
)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QPixmap

from models.cliente import Cliente
from ui.themes import obtener_tema, estilo_boton, estilo_input, estilo_tabla
from ui.iconos.iconos import iconos_crud, iconos_formulario, favicon

CARPETA_IMAGENES = os.path.join(os.path.dirname(__file__), "..", "..", "assets", "clientes")
os.makedirs(CARPETA_IMAGENES, exist_ok=True)


class ClientesWidget(QWidget):

    def __init__(self, repo):
        super().__init__()
        self.repo = repo
        self._construir_ui()
        self._aplicar_tema()
        self._cargar_datos()

    def _construir_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(16, 16, 16, 16)
        self.setLayout(layout)

        self.titulo = QLabel("Clientes")
        self.titulo.setStyleSheet("font-size: 22px; font-weight: bold; padding: 4px 0;")
        layout.addWidget(self.titulo)

        barra = QHBoxLayout()
        self.btn_nuevo    = QPushButton("  Nuevo")
        self.btn_editar   = QPushButton("  Editar")
        self.btn_eliminar = QPushButton("  Eliminar")
        self.btn_excel    = QPushButton("  Exportar Excel")
        self.btn_pdf      = QPushButton("  Exportar PDF")
        for btn in [self.btn_nuevo, self.btn_editar, self.btn_eliminar,
                    self.btn_excel, self.btn_pdf]:
            btn.setMinimumHeight(34)
        barra.addWidget(self.btn_nuevo)
        barra.addWidget(self.btn_editar)
        barra.addWidget(self.btn_eliminar)
        barra.addSpacing(20)
        barra.addWidget(self.btn_excel)
        barra.addWidget(self.btn_pdf)
        barra.addStretch()
        layout.addLayout(barra)

        filtros = QHBoxLayout()
        self.campo_busqueda = QLineEdit()
        self.campo_busqueda.setPlaceholderText("Buscar por nombre, contacto o RUC...")
        self.campo_busqueda.setMinimumHeight(32)
        self.filtro_tipo = QComboBox()
        self.filtro_tipo.addItems(["Todos los tipos", "PYME", "corporativo", "gobierno"])
        self.filtro_tipo.setMinimumHeight(32)
        self.filtro_clasif = QComboBox()
        self.filtro_clasif.addItems(["Toda clasificacion", "alto", "medio", "bajo"])
        self.filtro_clasif.setMinimumHeight(32)
        self.btn_buscar  = QPushButton("  Buscar")
        self.btn_limpiar = QPushButton("  Limpiar")
        self.btn_buscar.setMinimumHeight(32)
        self.btn_limpiar.setMinimumHeight(32)
        filtros.addWidget(self.campo_busqueda, 3)
        filtros.addWidget(self.filtro_tipo, 1)
        filtros.addWidget(self.filtro_clasif, 1)
        filtros.addWidget(self.btn_buscar)
        filtros.addWidget(self.btn_limpiar)
        layout.addLayout(filtros)

        self.tabla = QTableWidget()
        columnas = [
            "Codigo", "Tipo", "Razon Social", "Sector", "RUC",
            "Direccion", "Telefono", "Sitio Web", "Contacto",
            "Cargo", "Correo", "Tel. Directo", "Fecha 1ra Relacion",
            "Origen", "Clasificacion"
        ]
        self.tabla.setColumnCount(len(columnas))
        self.tabla.setHorizontalHeaderLabels(columnas)
        self.tabla.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.tabla.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.tabla.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.tabla.horizontalHeader().setStretchLastSection(True)
        self.tabla.verticalHeader().setVisible(False)
        self.tabla.setAlternatingRowColors(True)
        layout.addWidget(self.tabla)

        self.lbl_total = QLabel("Total: 0 registros")
        layout.addWidget(self.lbl_total)

        self.btn_nuevo.clicked.connect(self._abrir_formulario_nuevo)
        self.btn_editar.clicked.connect(self._abrir_formulario_editar)
        self.btn_eliminar.clicked.connect(self._eliminar)
        self.btn_excel.clicked.connect(self._exportar_excel)
        self.btn_pdf.clicked.connect(self._exportar_pdf)
        self.btn_buscar.clicked.connect(self._buscar)
        self.btn_limpiar.clicked.connect(self._limpiar)
        self.campo_busqueda.returnPressed.connect(self._buscar)
        self.filtro_tipo.currentIndexChanged.connect(self._buscar)
        self.filtro_clasif.currentIndexChanged.connect(self._buscar)

        iconos_crud(
            self.btn_nuevo, self.btn_editar, self.btn_eliminar,
            self.btn_excel, self.btn_pdf, self.btn_buscar, self.btn_limpiar
        )

    def _aplicar_tema(self):
        t = obtener_tema()
        self.setStyleSheet(f"background-color: {t['fondo']}; color: {t['texto']};")
        self.titulo.setStyleSheet(f"font-size: 22px; font-weight: bold; color: {t['titulo']};")
        self.lbl_total.setStyleSheet(f"font-size: 11px; color: {t['texto_suave']};")
        self.btn_nuevo.setStyleSheet(estilo_boton(t["btn_primario"],  t["btn_primario_txt"]))
        self.btn_editar.setStyleSheet(estilo_boton(t["btn_editar"],   t["btn_editar_txt"]))
        self.btn_eliminar.setStyleSheet(estilo_boton(t["btn_peligro"],t["btn_peligro_txt"]))
        self.btn_excel.setStyleSheet(estilo_boton(t["btn_excel"],     t["btn_excel_txt"]))
        self.btn_pdf.setStyleSheet(estilo_boton(t["btn_pdf"],         t["btn_pdf_txt"]))
        self.btn_buscar.setStyleSheet(estilo_boton(t["btn_buscar"],   t["btn_buscar_txt"]))
        self.btn_limpiar.setStyleSheet(estilo_boton(t["btn_limpiar"], t["btn_limpiar_txt"]))
        self.campo_busqueda.setStyleSheet(estilo_input(t))
        self.filtro_tipo.setStyleSheet(estilo_input(t))
        self.filtro_clasif.setStyleSheet(estilo_input(t))
        self.tabla.setStyleSheet(estilo_tabla(t))

    def _cargar_datos(self, clientes=None):
        if clientes is None:
            clientes = self.repo.select_all()
        self.tabla.setRowCount(len(clientes))
        for fila, c in enumerate(clientes):
            self.tabla.setItem(fila, 0,  QTableWidgetItem(str(c.codigo or "")))
            self.tabla.setItem(fila, 1,  QTableWidgetItem(str(c.tipo or "")))
            self.tabla.setItem(fila, 2,  QTableWidgetItem(str(c.razon_social or "")))
            self.tabla.setItem(fila, 3,  QTableWidgetItem(str(c.sector_actividad or "")))
            self.tabla.setItem(fila, 4,  QTableWidgetItem(str(c.ruc or "")))
            self.tabla.setItem(fila, 5,  QTableWidgetItem(str(c.direccion or "")))
            self.tabla.setItem(fila, 6,  QTableWidgetItem(str(c.telefono or "")))
            self.tabla.setItem(fila, 7,  QTableWidgetItem(str(c.sitio_web or "")))
            self.tabla.setItem(fila, 8,  QTableWidgetItem(str(c.contacto_principal or "")))
            self.tabla.setItem(fila, 9,  QTableWidgetItem(str(c.cargo_contacto or "")))
            self.tabla.setItem(fila, 10, QTableWidgetItem(str(c.correo_electronico or "")))
            self.tabla.setItem(fila, 11, QTableWidgetItem(str(c.telefono_directo or "")))
            self.tabla.setItem(fila, 12, QTableWidgetItem(str(c.fecha_primera_relacion or "")))
            self.tabla.setItem(fila, 13, QTableWidgetItem(str(c.origen_contacto or "")))
            self.tabla.setItem(fila, 14, QTableWidgetItem(str(c.clasificacion_potencial or "")))
        self.tabla.resizeColumnsToContents()
        self.lbl_total.setText(f"Total: {len(clientes)} registros")

    def _fila_seleccionada(self):
        fila = self.tabla.currentRow()
        if fila == -1:
            return None
        return self.tabla.item(fila, 0).text()

    def _buscar(self):
        termino    = self.campo_busqueda.text().strip()
        tipo       = self.filtro_tipo.currentText()
        clasif     = self.filtro_clasif.currentText()
        resultados = self.repo.search(termino) if termino else self.repo.select_all()
        if tipo   != "Todos los tipos":
            resultados = [c for c in resultados if c.tipo == tipo]
        if clasif != "Toda clasificacion":
            resultados = [c for c in resultados if c.clasificacion_potencial == clasif]
        self._cargar_datos(resultados)

    def _limpiar(self):
        self.campo_busqueda.clear()
        self.filtro_tipo.setCurrentIndex(0)
        self.filtro_clasif.setCurrentIndex(0)
        self._cargar_datos()

    def _abrir_formulario_nuevo(self):
        dialogo = FormularioCliente(self.repo)
        if dialogo.exec() == QDialog.DialogCode.Accepted:
            self._cargar_datos()

    def _abrir_formulario_editar(self):
        codigo = self._fila_seleccionada()
        if not codigo:
            QMessageBox.warning(self, "Aviso", "Selecciona un cliente para editar.")
            return
        cliente = self.repo.select_WHERE_codigo(codigo)
        dialogo = FormularioCliente(self.repo, cliente)
        if dialogo.exec() == QDialog.DialogCode.Accepted:
            self._cargar_datos()

    def _eliminar(self):
        codigo = self._fila_seleccionada()
        if not codigo:
            QMessageBox.warning(self, "Aviso", "Selecciona un cliente para eliminar.")
            return
        respuesta = QMessageBox.question(
            self, "Confirmar eliminacion",
            f"Estas seguro de eliminar el cliente {codigo}?\n\nEsta accion no se puede deshacer.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if respuesta == QMessageBox.StandardButton.Yes:
            try:
                self.repo.delete_cliente(codigo)
                for ext in [".jpg", ".png", ".gif"]:
                    ruta = os.path.join(CARPETA_IMAGENES, f"{codigo}{ext}")
                    if os.path.exists(ruta):
                        os.remove(ruta)
                self._cargar_datos()
                QMessageBox.information(self, "Exito", "Cliente eliminado correctamente.")
            except Exception as e:
                QMessageBox.critical(self, "No se puede eliminar", str(e))

    def _exportar_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            QMessageBox.critical(self, "Error", "Instala openpyxl:\npip install openpyxl")
            return
        dialogo = DialogoFiltroExportacion(self)
        if dialogo.exec() != QDialog.DialogCode.Accepted:
            return
        clientes = self._aplicar_filtros_exportacion(dialogo.obtener_filtros())
        ruta, _  = QFileDialog.getSaveFileName(self, "Guardar Excel", "clientes.xlsx", "Excel (*.xlsx)")
        if not ruta:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clientes"
        fill_enc = PatternFill("solid", fgColor="2C3E50")
        font_enc = Font(bold=True, color="FFFFFF", size=11)
        font_tit = Font(bold=True, size=14, color="2C3E50")
        borde    = Border(left=Side(style="thin"), right=Side(style="thin"),
                          top=Side(style="thin"),  bottom=Side(style="thin"))
        fill_par = PatternFill("solid", fgColor="EBF5FB")
        ws.merge_cells("A1:O1")
        ws["A1"] = "Reporte de Clientes"
        ws["A1"].font      = font_tit
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:O2")
        ws["A2"] = f"Generado: {date.today().strftime('%d/%m/%Y')}  |  Total: {len(clientes)} registros"
        ws["A2"].alignment = Alignment(horizontal="center")
        ws["A2"].font      = Font(italic=True, color="666666")
        cols = ["Codigo","Tipo","Razon Social","Sector","RUC","Direccion","Telefono",
                "Sitio Web","Contacto","Cargo","Correo","Tel. Directo",
                "Fecha 1ra Relacion","Origen","Clasificacion"]
        for i, col in enumerate(cols, 1):
            cell           = ws.cell(row=4, column=i, value=col)
            cell.font      = font_enc
            cell.fill      = fill_enc
            cell.alignment = Alignment(horizontal="center")
            cell.border    = borde
        for idx, c in enumerate(clientes, 5):
            datos = [c.codigo, c.tipo, c.razon_social, c.sector_actividad, c.ruc,
                     c.direccion, c.telefono, c.sitio_web, c.contacto_principal,
                     c.cargo_contacto, c.correo_electronico, c.telefono_directo,
                     str(c.fecha_primera_relacion or ""), c.origen_contacto,
                     c.clasificacion_potencial]
            fill = fill_par if idx % 2 == 0 else None
            for j, val in enumerate(datos, 1):
                cell        = ws.cell(row=idx, column=j, value=val or "")
                cell.border = borde
                if fill:
                    cell.fill = fill
        anchos = [12,12,30,20,15,20,15,25,25,20,30,15,18,15,14]
        for i, ancho in enumerate(anchos, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
        wb.save(ruta)
        QMessageBox.information(self, "Exportado", f"Excel guardado en:\n{ruta}")

    def _exportar_pdf(self):
        try:
            from reportlab.lib.pagesizes import landscape, A3
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import ParagraphStyle
            from reportlab.lib.units import cm
        except ImportError:
            QMessageBox.critical(self, "Error", "Instala reportlab:\npip install reportlab")
            return
        dialogo = DialogoFiltroExportacion(self)
        if dialogo.exec() != QDialog.DialogCode.Accepted:
            return
        clientes = self._aplicar_filtros_exportacion(dialogo.obtener_filtros())
        ruta, _  = QFileDialog.getSaveFileName(self, "Guardar PDF", "clientes.pdf", "PDF (*.pdf)")
        if not ruta:
            return
        doc   = SimpleDocTemplate(ruta, pagesize=landscape(A3),
                                  topMargin=1.5*cm, bottomMargin=1.5*cm)
        story = []
        st = ParagraphStyle("t", fontSize=16, fontName="Helvetica-Bold",
                             spaceAfter=4, textColor=colors.HexColor("#2C3E50"))
        ss = ParagraphStyle("s", fontSize=9,  fontName="Helvetica",
                             spaceAfter=12, textColor=colors.grey)
        story.append(Paragraph("Reporte de Clientes", st))
        story.append(Paragraph(
            f"Generado: {date.today().strftime('%d/%m/%Y')}  |  Total: {len(clientes)} registros", ss))
        story.append(Spacer(1, 0.3*cm))
        enc   = ["Codigo","Tipo","Razon Social","Sector","RUC",
                 "Contacto","Correo","Clasificacion"]
        datos = [enc]
        for c in clientes:
            datos.append([c.codigo or "", c.tipo or "", c.razon_social or "",
                          c.sector_actividad or "", c.ruc or "",
                          c.contacto_principal or "", c.correo_electronico or "",
                          c.clasificacion_potencial or ""])
        tabla = Table(datos, repeatRows=1)
        tabla.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),(-1,0),  colors.HexColor("#2C3E50")),
            ("TEXTCOLOR",     (0,0),(-1,0),  colors.white),
            ("FONTNAME",      (0,0),(-1,0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0,0),(-1,-1), 8),
            ("ALIGN",         (0,0),(-1,-1), "CENTER"),
            ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, colors.HexColor("#EBF5FB")]),
            ("GRID",          (0,0),(-1,-1), 0.5, colors.HexColor("#CCCCCC")),
            ("TOPPADDING",    (0,0),(-1,-1), 5),
            ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ]))
        story.append(tabla)
        doc.build(story)
        QMessageBox.information(self, "Exportado", f"PDF guardado en:\n{ruta}")

    def _aplicar_filtros_exportacion(self, filtros):
        tipo   = filtros.get("tipo")
        clasif = filtros.get("clasificacion")
        datos  = self.repo.select_all()
        if tipo:
            datos = [c for c in datos if c.tipo == tipo]
        if clasif:
            datos = [c for c in datos if c.clasificacion_potencial == clasif]
        return datos


class DialogoFiltroExportacion(QDialog):

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Filtros de exportacion")
        self.setWindowIcon(favicon("clientes"))
        self.setMinimumWidth(320)
        t = obtener_tema()
        self.setStyleSheet(f"background-color: {t['fondo_panel']}; color: {t['texto']};")
        layout = QVBoxLayout()
        self.setLayout(layout)
        layout.addWidget(QLabel("Filtrar por tipo:"))
        self.combo_tipo = QComboBox()
        self.combo_tipo.addItems(["Todos", "PYME", "corporativo", "gobierno"])
        self.combo_tipo.setStyleSheet(estilo_input(t))
        layout.addWidget(self.combo_tipo)
        layout.addWidget(QLabel("Filtrar por clasificacion:"))
        self.combo_clasif = QComboBox()
        self.combo_clasif.addItems(["Todos", "alto", "medio", "bajo"])
        self.combo_clasif.setStyleSheet(estilo_input(t))
        layout.addWidget(self.combo_clasif)
        barra      = QHBoxLayout()
        btn_ok     = QPushButton("  Exportar")
        btn_cancel = QPushButton("  Cancelar")
        btn_ok.setStyleSheet(estilo_boton(t["btn_primario"], t["btn_primario_txt"]))
        btn_cancel.setStyleSheet(estilo_boton(t["btn_limpiar"], t["btn_limpiar_txt"]))
        barra.addStretch()
        barra.addWidget(btn_ok)
        barra.addWidget(btn_cancel)
        layout.addLayout(barra)
        btn_ok.clicked.connect(self.accept)
        btn_cancel.clicked.connect(self.reject)

    def obtener_filtros(self):
        tipo   = self.combo_tipo.currentText()
        clasif = self.combo_clasif.currentText()
        return {
            "tipo":          None if tipo   == "Todos" else tipo,
            "clasificacion": None if clasif == "Todos" else clasif,
        }


class FormularioCliente(QDialog):

    def __init__(self, repo, cliente=None):
        super().__init__()
        self.repo        = repo
        self.cliente     = cliente
        self.editando    = cliente is not None
        self.ruta_imagen = None
        t                = obtener_tema()

        self.setWindowTitle("Editar Cliente" if self.editando else "Nuevo Cliente")
        self.setWindowIcon(favicon("clientes"))
        self.setMinimumWidth(560)
        self.setStyleSheet(f"background-color: {t['fondo_panel']}; color: {t['texto']};")

        layout_principal = QHBoxLayout()
        self.setLayout(layout_principal)

        scroll     = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("border: none;")
        contenedor = QWidget()
        scroll.setWidget(contenedor)
        form_layout = QVBoxLayout()
        contenedor.setLayout(form_layout)

        form = QFormLayout()
        form.setSpacing(10)
        estilo = estilo_input(t)

        self.campo_codigo    = QLineEdit()
        self.campo_tipo      = QComboBox()
        self.campo_tipo.addItems(["PYME", "corporativo", "gobierno"])
        self.campo_razon     = QLineEdit()
        self.campo_sector    = QLineEdit()
        self.campo_ruc       = QLineEdit()
        self.campo_direccion = QLineEdit()
        self.campo_telefono  = QLineEdit()
        self.campo_sitio_web = QLineEdit()
        self.campo_contacto  = QLineEdit()
        self.campo_cargo     = QLineEdit()
        self.campo_correo    = QLineEdit()
        self.campo_tel_dir   = QLineEdit()
        self.campo_fecha     = QDateEdit()
        self.campo_fecha.setCalendarPopup(True)
        self.campo_fecha.setDate(QDate.currentDate())
        self.campo_fecha.setDisplayFormat("dd/MM/yyyy")
        self.campo_origen    = QLineEdit()
        self.campo_clasif    = QComboBox()
        self.campo_clasif.addItems(["", "alto", "medio", "bajo"])

        for w in [self.campo_codigo, self.campo_razon, self.campo_sector,
                  self.campo_ruc, self.campo_direccion, self.campo_telefono,
                  self.campo_sitio_web, self.campo_contacto, self.campo_cargo,
                  self.campo_correo, self.campo_tel_dir, self.campo_origen,
                  self.campo_tipo, self.campo_clasif, self.campo_fecha]:
            w.setStyleSheet(estilo)

        form.addRow("Codigo *:",           self.campo_codigo)
        form.addRow("Tipo *:",             self.campo_tipo)
        form.addRow("Razon Social *:",     self.campo_razon)
        form.addRow("Sector:",             self.campo_sector)
        form.addRow("RUC:",                self.campo_ruc)
        form.addRow("Direccion:",          self.campo_direccion)
        form.addRow("Telefono:",           self.campo_telefono)
        form.addRow("Sitio Web:",          self.campo_sitio_web)
        form.addRow("Contacto:",           self.campo_contacto)
        form.addRow("Cargo:",              self.campo_cargo)
        form.addRow("Correo:",             self.campo_correo)
        form.addRow("Tel. Directo:",       self.campo_tel_dir)
        form.addRow("Fecha 1ra Relacion:", self.campo_fecha)
        form.addRow("Origen:",             self.campo_origen)
        form.addRow("Clasificacion:",      self.campo_clasif)
        form_layout.addLayout(form)

        barra = QHBoxLayout()
        self.btn_guardar  = QPushButton("  Guardar")
        self.btn_cancelar = QPushButton("  Cancelar")
        self.btn_guardar.setMinimumHeight(34)
        self.btn_cancelar.setMinimumHeight(34)
        self.btn_guardar.setStyleSheet(estilo_boton(t["btn_primario"],  t["btn_primario_txt"]))
        self.btn_cancelar.setStyleSheet(estilo_boton(t["btn_limpiar"],  t["btn_limpiar_txt"]))
        barra.addStretch()
        barra.addWidget(self.btn_guardar)
        barra.addWidget(self.btn_cancelar)
        form_layout.addLayout(barra)

        self.btn_guardar.clicked.connect(self._guardar)
        self.btn_cancelar.clicked.connect(self.reject)
        layout_principal.addWidget(scroll, 2)

        panel_imagen = QFrame()
        panel_imagen.setFixedWidth(180)
        panel_imagen.setStyleSheet(
            f"background-color: {t['fondo_input']}; "
            f"border-radius: 8px; border: 1px solid {t['borde']};"
        )
        img_layout = QVBoxLayout()
        img_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        panel_imagen.setLayout(img_layout)

        lbl_img_titulo = QLabel("Imagen del Cliente")
        lbl_img_titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_img_titulo.setStyleSheet(
            f"color: {t['texto_suave']}; font-size: 11px; font-weight: bold; border: none;"
        )
        img_layout.addWidget(lbl_img_titulo)

        self.lbl_imagen = QLabel()
        self.lbl_imagen.setFixedSize(150, 150)
        self.lbl_imagen.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_imagen.setStyleSheet(
            f"border: 2px dashed {t['borde']}; border-radius: 6px; color: {t['texto_suave']};"
        )
        self.lbl_imagen.setText("Sin imagen")
        img_layout.addWidget(self.lbl_imagen)

        self.btn_cargar       = QPushButton("  Cargar imagen")
        self.btn_eliminar_img = QPushButton("  Quitar imagen")
        self.btn_cargar.setStyleSheet(estilo_boton(t["btn_buscar"],  t["btn_buscar_txt"]))
        self.btn_eliminar_img.setStyleSheet(estilo_boton(t["btn_peligro"], t["btn_peligro_txt"]))
        img_layout.addWidget(self.btn_cargar)
        img_layout.addWidget(self.btn_eliminar_img)

        lbl_formatos = QLabel("JPG, PNG, GIF\nMax: 5MB")
        lbl_formatos.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_formatos.setStyleSheet(
            f"color: {t['texto_suave']}; font-size: 10px; border: none;"
        )
        img_layout.addWidget(lbl_formatos)
        img_layout.addStretch()

        self.btn_cargar.clicked.connect(self._cargar_imagen)
        self.btn_eliminar_img.clicked.connect(self._quitar_imagen)
        layout_principal.addWidget(panel_imagen, 1)

        iconos_formulario(self.btn_guardar, self.btn_cancelar, self.btn_cargar)

        if self.editando:
            self.campo_codigo.setText(cliente.codigo)
            self.campo_codigo.setEnabled(False)
            self.campo_tipo.setCurrentText(cliente.tipo or "")
            self.campo_razon.setText(cliente.razon_social or "")
            self.campo_sector.setText(cliente.sector_actividad or "")
            self.campo_ruc.setText(cliente.ruc or "")
            self.campo_direccion.setText(cliente.direccion or "")
            self.campo_telefono.setText(cliente.telefono or "")
            self.campo_sitio_web.setText(cliente.sitio_web or "")
            self.campo_contacto.setText(cliente.contacto_principal or "")
            self.campo_cargo.setText(cliente.cargo_contacto or "")
            self.campo_correo.setText(cliente.correo_electronico or "")
            self.campo_tel_dir.setText(cliente.telefono_directo or "")
            if cliente.fecha_primera_relacion:
                f = cliente.fecha_primera_relacion
                self.campo_fecha.setDate(QDate(f.year, f.month, f.day))
            self.campo_origen.setText(cliente.origen_contacto or "")
            self.campo_clasif.setCurrentText(cliente.clasificacion_potencial or "")
            self._cargar_imagen_existente(cliente.codigo)

    def _cargar_imagen_existente(self, codigo):
        for ext in [".jpg", ".png", ".gif"]:
            ruta = os.path.join(CARPETA_IMAGENES, f"{codigo}{ext}")
            if os.path.exists(ruta):
                self._mostrar_imagen(ruta)
                self.ruta_imagen = ruta
                break

    def _cargar_imagen(self):
        ruta, _ = QFileDialog.getOpenFileName(
            self, "Seleccionar imagen", "",
            "Imagenes (*.jpg *.jpeg *.png *.gif)"
        )
        if not ruta:
            return
        if os.path.getsize(ruta) > 5 * 1024 * 1024:
            QMessageBox.warning(self, "Error", "La imagen supera el tamaño maximo de 5MB.")
            return
        try:
            from PIL import Image
            img = Image.open(ruta)
            if img.format not in ["JPEG", "PNG", "GIF"]:
                QMessageBox.warning(self, "Error", "Formato no soportado. Usa JPG, PNG o GIF.")
                return
        except Exception:
            QMessageBox.warning(self, "Error", "No se pudo leer la imagen.")
            return
        self.ruta_imagen = ruta
        self._mostrar_imagen(ruta)

    def _mostrar_imagen(self, ruta):
        pixmap = QPixmap(ruta).scaled(
            150, 150,
            Qt.AspectRatioMode.KeepAspectRatio,
            Qt.TransformationMode.SmoothTransformation
        )
        self.lbl_imagen.setPixmap(pixmap)
        self.lbl_imagen.setText("")

    def _quitar_imagen(self):
        self.ruta_imagen = None
        self.lbl_imagen.setPixmap(QPixmap())
        self.lbl_imagen.setText("Sin imagen")

    def _validar(self):
        codigo = self.campo_codigo.text().strip()
        razon  = self.campo_razon.text().strip()
        correo = self.campo_correo.text().strip()
        tel    = self.campo_telefono.text().strip()
        if not codigo:
            QMessageBox.warning(self, "Error", "El codigo es obligatorio.")
            return False
        if len(codigo) < 3:
            QMessageBox.warning(self, "Error", "El codigo debe tener al menos 3 caracteres.")
            return False
        if not razon:
            QMessageBox.warning(self, "Error", "La razon social es obligatoria.")
            return False
        if len(razon) < 3:
            QMessageBox.warning(self, "Error", "La razon social debe tener al menos 3 caracteres.")
            return False
        if correo and not re.match(r"^[\w\.-]+@[\w\.-]+\.\w{2,}$", correo):
            QMessageBox.warning(self, "Error", "El formato del correo no es valido.")
            return False
        if tel and not re.match(r"^[\d\+\-\(\)\s]+$", tel):
            QMessageBox.warning(self, "Error", "El telefono solo puede contener numeros y + - ( )")
            return False
        return True

    def _guardar(self):
        if not self._validar():
            return
        accion = "actualizar" if self.editando else "guardar"
        respuesta = QMessageBox.question(
            self, "Confirmar",
            f"Deseas {accion} este cliente?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if respuesta != QMessageBox.StandardButton.Yes:
            return
        fecha_q  = self.campo_fecha.date()
        fecha_py = date(fecha_q.year(), fecha_q.month(), fecha_q.day())
        cliente  = Cliente(
            codigo                  = self.campo_codigo.text().strip(),
            tipo                    = self.campo_tipo.currentText(),
            razon_social            = self.campo_razon.text().strip(),
            sector_actividad        = self.campo_sector.text().strip() or None,
            ruc                     = self.campo_ruc.text().strip() or None,
            direccion               = self.campo_direccion.text().strip() or None,
            telefono                = self.campo_telefono.text().strip() or None,
            sitio_web               = self.campo_sitio_web.text().strip() or None,
            contacto_principal      = self.campo_contacto.text().strip() or None,
            cargo_contacto          = self.campo_cargo.text().strip() or None,
            correo_electronico      = self.campo_correo.text().strip() or None,
            telefono_directo        = self.campo_tel_dir.text().strip() or None,
            fecha_primera_relacion  = fecha_py,
            origen_contacto         = self.campo_origen.text().strip() or None,
            clasificacion_potencial = self.campo_clasif.currentText() or None
        )
        try:
            if self.editando:
                self.repo.update(cliente)
            else:
                self.repo.insert(cliente)
            if self.ruta_imagen:
                ext     = os.path.splitext(self.ruta_imagen)[1].lower()
                destino = os.path.join(CARPETA_IMAGENES, f"{cliente.codigo}{ext}")
                if self.ruta_imagen != destino:
                    shutil.copy2(self.ruta_imagen, destino)
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Error al guardar", str(e))
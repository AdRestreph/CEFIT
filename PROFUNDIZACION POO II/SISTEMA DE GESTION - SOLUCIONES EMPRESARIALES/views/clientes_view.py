import re
from datetime import date

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QTableWidget,
    QTableWidgetItem, QDialog, QFormLayout, QLineEdit, QComboBox,
    QLabel, QMessageBox, QHeaderView, QDateEdit, QFileDialog,
    QScrollArea, QFrame
)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QPixmap

from ui.themes import obtener_tema, estilo_boton, estilo_input, estilo_tabla
from ui.iconos.iconos import iconos_crud, iconos_formulario, favicon
from views.dialogs.filtro_exportacion_dialog import DialogoFiltroExportacion


class ClientesView(QWidget):
    """
    Vista del modulo de Clientes.
    Solo se encarga de mostrar datos y capturar eventos del usuario.
    Toda la logica de negocio se delega al ClienteController.
    """

    def __init__(self, controller):
        super().__init__()
        self.controller = controller
        self._construir_ui()
        self._aplicar_tema()
        self._cargar_datos()

    # ------------------------------------------------------------------ UI

    def _construir_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(16, 16, 16, 16)
        self.setLayout(layout)

        self.titulo = QLabel("Clientes")
        layout.addWidget(self.titulo)

        barra = QHBoxLayout()
        self.btn_nuevo    = QPushButton("  Nuevo")
        self.btn_editar   = QPushButton("  Editar")
        self.btn_eliminar = QPushButton("  Eliminar")
        self.btn_excel    = QPushButton("  Exportar Excel")
        self.btn_pdf      = QPushButton("  Exportar PDF")
        for btn in [self.btn_nuevo, self.btn_editar, self.btn_eliminar,
                    self.btn_excel, self.btn_pdf]:
            btn.setMinimumHeight(34)
        barra.addWidget(self.btn_nuevo)
        barra.addWidget(self.btn_editar)
        barra.addWidget(self.btn_eliminar)
        barra.addSpacing(20)
        barra.addWidget(self.btn_excel)
        barra.addWidget(self.btn_pdf)
        barra.addStretch()
        layout.addLayout(barra)

        filtros = QHBoxLayout()
        self.campo_busqueda = QLineEdit()
        self.campo_busqueda.setPlaceholderText("Buscar por nombre, contacto o RUC...")
        self.campo_busqueda.setMinimumHeight(32)
        self.filtro_tipo = QComboBox()
        self.filtro_tipo.addItems(["Todos los tipos", "PYME", "corporativo", "gobierno"])
        self.filtro_tipo.setMinimumHeight(32)
        self.filtro_clasif = QComboBox()
        self.filtro_clasif.addItems(["Toda clasificacion", "alto", "medio", "bajo"])
        self.filtro_clasif.setMinimumHeight(32)
        self.btn_buscar  = QPushButton("  Buscar")
        self.btn_limpiar = QPushButton("  Limpiar")
        self.btn_buscar.setMinimumHeight(32)
        self.btn_limpiar.setMinimumHeight(32)
        filtros.addWidget(self.campo_busqueda, 3)
        filtros.addWidget(self.filtro_tipo, 1)
        filtros.addWidget(self.filtro_clasif, 1)
        filtros.addWidget(self.btn_buscar)
        filtros.addWidget(self.btn_limpiar)
        layout.addLayout(filtros)

        self.tabla = QTableWidget()
        columnas = [
            "Codigo", "Tipo", "Razon Social", "Sector", "RUC",
            "Direccion", "Telefono", "Sitio Web", "Contacto",
            "Cargo", "Correo", "Tel. Directo", "Fecha 1ra Relacion",
            "Origen", "Clasificacion"
        ]
        self.tabla.setColumnCount(len(columnas))
        self.tabla.setHorizontalHeaderLabels(columnas)
        self.tabla.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.tabla.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.tabla.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.tabla.horizontalHeader().setStretchLastSection(True)
        self.tabla.verticalHeader().setVisible(False)
        self.tabla.setAlternatingRowColors(True)
        layout.addWidget(self.tabla)

        self.lbl_total = QLabel("Total: 0 registros")
        layout.addWidget(self.lbl_total)

        # Conexiones
        self.btn_nuevo.clicked.connect(self._abrir_formulario_nuevo)
        self.btn_editar.clicked.connect(self._abrir_formulario_editar)
        self.btn_eliminar.clicked.connect(self._eliminar)
        self.btn_excel.clicked.connect(self._exportar_excel)
        self.btn_pdf.clicked.connect(self._exportar_pdf)
        self.btn_buscar.clicked.connect(self._buscar)
        self.btn_limpiar.clicked.connect(self._limpiar)
        self.campo_busqueda.returnPressed.connect(self._buscar)
        self.filtro_tipo.currentIndexChanged.connect(self._buscar)
        self.filtro_clasif.currentIndexChanged.connect(self._buscar)

        iconos_crud(
            self.btn_nuevo, self.btn_editar, self.btn_eliminar,
            self.btn_excel, self.btn_pdf, self.btn_buscar, self.btn_limpiar
        )

    def _aplicar_tema(self):
        t = obtener_tema()
        self.setStyleSheet(f"background-color: {t['fondo']}; color: {t['texto']};")
        self.titulo.setStyleSheet(f"font-size: 22px; font-weight: bold; color: {t['titulo']};")
        self.lbl_total.setStyleSheet(f"font-size: 11px; color: {t['texto_suave']};")
        self.btn_nuevo.setStyleSheet(estilo_boton(t["btn_primario"],  t["btn_primario_txt"]))
        self.btn_editar.setStyleSheet(estilo_boton(t["btn_editar"],   t["btn_editar_txt"]))
        self.btn_eliminar.setStyleSheet(estilo_boton(t["btn_peligro"], t["btn_peligro_txt"]))
        self.btn_excel.setStyleSheet(estilo_boton(t["btn_excel"],     t["btn_excel_txt"]))
        self.btn_pdf.setStyleSheet(estilo_boton(t["btn_pdf"],         t["btn_pdf_txt"]))
        self.btn_buscar.setStyleSheet(estilo_boton(t["btn_buscar"],   t["btn_buscar_txt"]))
        self.btn_limpiar.setStyleSheet(estilo_boton(t["btn_limpiar"], t["btn_limpiar_txt"]))
        self.campo_busqueda.setStyleSheet(estilo_input(t))
        self.filtro_tipo.setStyleSheet(estilo_input(t))
        self.filtro_clasif.setStyleSheet(estilo_input(t))
        self.tabla.setStyleSheet(estilo_tabla(t))

    # --------------------------------------------------------------- Datos

    def _cargar_datos(self, clientes=None):
        if clientes is None:
            clientes = self.controller.obtener_todos()
        self.tabla.setRowCount(len(clientes))
        for fila, c in enumerate(clientes):
            self.tabla.setItem(fila, 0,  QTableWidgetItem(str(c.codigo or "")))
            self.tabla.setItem(fila, 1,  QTableWidgetItem(str(c.tipo or "")))
            self.tabla.setItem(fila, 2,  QTableWidgetItem(str(c.razon_social or "")))
            self.tabla.setItem(fila, 3,  QTableWidgetItem(str(c.sector_actividad or "")))
            self.tabla.setItem(fila, 4,  QTableWidgetItem(str(c.ruc or "")))
            self.tabla.setItem(fila, 5,  QTableWidgetItem(str(c.direccion or "")))
            self.tabla.setItem(fila, 6,  QTableWidgetItem(str(c.telefono or "")))
            self.tabla.setItem(fila, 7,  QTableWidgetItem(str(c.sitio_web or "")))
            self.tabla.setItem(fila, 8,  QTableWidgetItem(str(c.contacto_principal or "")))
            self.tabla.setItem(fila, 9,  QTableWidgetItem(str(c.cargo_contacto or "")))
            self.tabla.setItem(fila, 10, QTableWidgetItem(str(c.correo_electronico or "")))
            self.tabla.setItem(fila, 11, QTableWidgetItem(str(c.telefono_directo or "")))
            self.tabla.setItem(fila, 12, QTableWidgetItem(str(c.fecha_primera_relacion or "")))
            self.tabla.setItem(fila, 13, QTableWidgetItem(str(c.origen_contacto or "")))
            self.tabla.setItem(fila, 14, QTableWidgetItem(str(c.clasificacion_potencial or "")))
        self.tabla.resizeColumnsToContents()
        self.lbl_total.setText(f"Total: {len(clientes)} registros")

    def _fila_seleccionada(self):
        fila = self.tabla.currentRow()
        if fila == -1:
            return None
        return self.tabla.item(fila, 0).text()

    # ------------------------------------------------------------ Acciones

    def _buscar(self):
        termino = self.campo_busqueda.text().strip()
        tipo    = self.filtro_tipo.currentText()
        clasif  = self.filtro_clasif.currentText()
        resultados = self.controller.buscar(
            termino=termino,
            tipo=None if tipo == "Todos los tipos" else tipo,
            clasificacion=None if clasif == "Toda clasificacion" else clasif,
        )
        self._cargar_datos(resultados)

    def _limpiar(self):
        self.campo_busqueda.clear()
        self.filtro_tipo.setCurrentIndex(0)
        self.filtro_clasif.setCurrentIndex(0)
        self._cargar_datos()

    def _abrir_formulario_nuevo(self):
        from views.dialogs.cliente_form_dialog import FormularioCliente
        dialogo = FormularioCliente(self.controller)
        if dialogo.exec() == QDialog.DialogCode.Accepted:
            self._cargar_datos()

    def _abrir_formulario_editar(self):
        codigo = self._fila_seleccionada()
        if not codigo:
            QMessageBox.warning(self, "Aviso", "Selecciona un cliente para editar.")
            return
        cliente = self.controller.obtener_por_codigo(codigo)
        from views.dialogs.cliente_form_dialog import FormularioCliente
        dialogo = FormularioCliente(self.controller, cliente)
        if dialogo.exec() == QDialog.DialogCode.Accepted:
            self._cargar_datos()

    def _eliminar(self):
        codigo = self._fila_seleccionada()
        if not codigo:
            QMessageBox.warning(self, "Aviso", "Selecciona un cliente para eliminar.")
            return
        respuesta = QMessageBox.question(
            self, "Confirmar eliminacion",
            f"Estas seguro de eliminar el cliente {codigo}?\n\nEsta accion no se puede deshacer.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if respuesta == QMessageBox.StandardButton.Yes:
            try:
                self.controller.eliminar(codigo)
                self._cargar_datos()
                QMessageBox.information(self, "Exito", "Cliente eliminado correctamente.")
            except Exception as e:
                QMessageBox.critical(self, "No se puede eliminar", str(e))

    # ---------------------------------------------------------- Exportacion

    def _exportar_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            QMessageBox.critical(self, "Error", "Instala openpyxl:\npip install openpyxl")
            return
        dialogo = DialogoFiltroExportacion(self)
        if dialogo.exec() != QDialog.DialogCode.Accepted:
            return
        filtros  = dialogo.obtener_filtros()
        clientes = self.controller.obtener_para_exportar(**filtros)
        ruta, _  = QFileDialog.getSaveFileName(self, "Guardar Excel", "clientes.xlsx", "Excel (*.xlsx)")
        if not ruta:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clientes"
        fill_enc = PatternFill("solid", fgColor="2C3E50")
        font_enc = Font(bold=True, color="FFFFFF", size=11)
        font_tit = Font(bold=True, size=14, color="2C3E50")
        borde    = Border(left=Side(style="thin"), right=Side(style="thin"),
                          top=Side(style="thin"),  bottom=Side(style="thin"))
        fill_par = PatternFill("solid", fgColor="EBF5FB")

        ws.merge_cells("A1:O1")
        ws["A1"] = "Reporte de Clientes"
        ws["A1"].font      = font_tit
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:O2")
        ws["A2"] = f"Generado: {date.today().strftime('%d/%m/%Y')}  |  Total: {len(clientes)} registros"
        ws["A2"].alignment = Alignment(horizontal="center")
        ws["A2"].font      = Font(italic=True, color="666666")

        cols = ["Codigo","Tipo","Razon Social","Sector","RUC","Direccion","Telefono",
                "Sitio Web","Contacto","Cargo","Correo","Tel. Directo",
                "Fecha 1ra Relacion","Origen","Clasificacion"]
        for i, col in enumerate(cols, 1):
            cell           = ws.cell(row=4, column=i, value=col)
            cell.font      = font_enc
            cell.fill      = fill_enc
            cell.alignment = Alignment(horizontal="center")
            cell.border    = borde

        for idx, c in enumerate(clientes, 5):
            datos = [c.codigo, c.tipo, c.razon_social, c.sector_actividad, c.ruc,
                     c.direccion, c.telefono, c.sitio_web, c.contacto_principal,
                     c.cargo_contacto, c.correo_electronico, c.telefono_directo,
                     str(c.fecha_primera_relacion or ""), c.origen_contacto,
                     c.clasificacion_potencial]
            fill = fill_par if idx % 2 == 0 else None
            for j, val in enumerate(datos, 1):
                cell        = ws.cell(row=idx, column=j, value=val or "")
                cell.border = borde
                if fill:
                    cell.fill = fill

        anchos = [12,12,30,20,15,20,15,25,25,20,30,15,18,15,14]
        for i, ancho in enumerate(anchos, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

        wb.save(ruta)
        QMessageBox.information(self, "Exportado", f"Excel guardado en:\n{ruta}")

    def _exportar_pdf(self):
        try:
            from reportlab.lib.pagesizes import landscape, A3
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import ParagraphStyle
            from reportlab.lib.units import cm
        except ImportError:
            QMessageBox.critical(self, "Error", "Instala reportlab:\npip install reportlab")
            return
        dialogo = DialogoFiltroExportacion(self)
        if dialogo.exec() != QDialog.DialogCode.Accepted:
            return
        filtros  = dialogo.obtener_filtros()
        clientes = self.controller.obtener_para_exportar(**filtros)
        ruta, _  = QFileDialog.getSaveFileName(self, "Guardar PDF", "clientes.pdf", "PDF (*.pdf)")
        if not ruta:
            return

        doc   = SimpleDocTemplate(ruta, pagesize=landscape(A3),
                                  topMargin=1.5*cm, bottomMargin=1.5*cm)
        story = []
        st = ParagraphStyle("t", fontSize=16, fontName="Helvetica-Bold",
                             spaceAfter=4, textColor=colors.HexColor("#2C3E50"))
        ss = ParagraphStyle("s", fontSize=9,  fontName="Helvetica",
                             spaceAfter=12, textColor=colors.grey)
        story.append(Paragraph("Reporte de Clientes", st))
        story.append(Paragraph(
            f"Generado: {date.today().strftime('%d/%m/%Y')}  |  Total: {len(clientes)} registros", ss))
        story.append(Spacer(1, 0.3*cm))

        enc   = ["Codigo","Tipo","Razon Social","Sector","RUC",
                 "Contacto","Correo","Clasificacion"]
        datos = [enc]
        for c in clientes:
            datos.append([c.codigo or "", c.tipo or "", c.razon_social or "",
                          c.sector_actividad or "", c.ruc or "",
                          c.contacto_principal or "", c.correo_electronico or "",
                          c.clasificacion_potencial or ""])
        tabla = Table(datos, repeatRows=1)
        tabla.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),(-1,0),  colors.HexColor("#2C3E50")),
            ("TEXTCOLOR",     (0,0),(-1,0),  colors.white),
            ("FONTNAME",      (0,0),(-1,0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0,0),(-1,-1), 8),
            ("ALIGN",         (0,0),(-1,-1), "CENTER"),
            ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, colors.HexColor("#EBF5FB")]),
            ("GRID",          (0,0),(-1,-1), 0.5, colors.HexColor("#CCCCCC")),
            ("TOPPADDING",    (0,0),(-1,-1), 5),
            ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ]))
        story.append(tabla)
        doc.build(story)
        QMessageBox.information(self, "Exportado", f"PDF guardado en:\n{ruta}")

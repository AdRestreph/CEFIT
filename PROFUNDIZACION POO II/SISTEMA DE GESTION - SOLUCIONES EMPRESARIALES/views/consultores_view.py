from datetime import date

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QTableWidget,
    QTableWidgetItem, QDialog, QFormLayout, QLineEdit, QComboBox,
    QLabel, QMessageBox, QSpinBox, QDoubleSpinBox, QFileDialog
)
from PyQt6.QtCore import Qt

from ui.themes import obtener_tema, estilo_boton, estilo_input, estilo_tabla
from ui.iconos.iconos import iconos_crud, iconos_formulario, favicon


class ConsultoresView(QWidget):
    """
    Vista del modulo de Consultores.
    Delega logica al ConsultorController.
    """

    def __init__(self, controller):
        super().__init__()
        self.controller = controller
        self._construir_ui()
        self._aplicar_tema()
        self._cargar_datos()

    def _construir_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(16, 16, 16, 16)
        self.setLayout(layout)

        self.titulo = QLabel("Consultores")
        layout.addWidget(self.titulo)

        barra = QHBoxLayout()
        self.btn_nuevo    = QPushButton("  Nuevo")
        self.btn_editar   = QPushButton("  Editar")
        self.btn_eliminar = QPushButton("  Eliminar")
        self.btn_excel    = QPushButton("  Exportar Excel")
        self.btn_pdf      = QPushButton("  Exportar PDF")
        for btn in [self.btn_nuevo, self.btn_editar, self.btn_eliminar,
                    self.btn_excel, self.btn_pdf]:
            btn.setMinimumHeight(34)
        barra.addWidget(self.btn_nuevo)
        barra.addWidget(self.btn_editar)
        barra.addWidget(self.btn_eliminar)
        barra.addSpacing(20)
        barra.addWidget(self.btn_excel)
        barra.addWidget(self.btn_pdf)
        barra.addStretch()
        layout.addLayout(barra)

        filtros = QHBoxLayout()
        self.campo_busqueda = QLineEdit()
        self.campo_busqueda.setPlaceholderText("Buscar por nombre o especialidad...")
        self.campo_busqueda.setMinimumHeight(32)
        self.filtro_nivel = QComboBox()
        self.filtro_nivel.addItems(["Todos los niveles", "junior", "senior", "gerente", "socio"])
        self.filtro_nivel.setMinimumHeight(32)
        self.filtro_disp = QComboBox()
        self.filtro_disp.addItems(["Toda disponibilidad", "tiempo completo", "medio tiempo"])
        self.filtro_disp.setMinimumHeight(32)
        self.btn_buscar  = QPushButton("  Buscar")
        self.btn_limpiar = QPushButton("  Limpiar")
        self.btn_buscar.setMinimumHeight(32)
        self.btn_limpiar.setMinimumHeight(32)
        filtros.addWidget(self.campo_busqueda, 3)
        filtros.addWidget(self.filtro_nivel, 1)
        filtros.addWidget(self.filtro_disp, 1)
        filtros.addWidget(self.btn_buscar)
        filtros.addWidget(self.btn_limpiar)
        layout.addLayout(filtros)

        self.tabla = QTableWidget()
        columnas = ["Codigo", "Nombres", "Apellidos", "Nivel",
                    "Especialidades", "Tarifa/Hora", "Idiomas", "Disponibilidad"]
        self.tabla.setColumnCount(len(columnas))
        self.tabla.setHorizontalHeaderLabels(columnas)
        self.tabla.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.tabla.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.tabla.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.tabla.horizontalHeader().setStretchLastSection(True)
        self.tabla.verticalHeader().setVisible(False)
        self.tabla.setAlternatingRowColors(True)
        layout.addWidget(self.tabla)

        self.lbl_total = QLabel("Total: 0 registros")
        layout.addWidget(self.lbl_total)

        self.btn_nuevo.clicked.connect(self._abrir_formulario_nuevo)
        self.btn_editar.clicked.connect(self._abrir_formulario_editar)
        self.btn_eliminar.clicked.connect(self._eliminar)
        self.btn_excel.clicked.connect(self._exportar_excel)
        self.btn_buscar.clicked.connect(self._buscar)
        self.btn_limpiar.clicked.connect(self._limpiar)
        self.campo_busqueda.returnPressed.connect(self._buscar)
        self.filtro_nivel.currentIndexChanged.connect(self._buscar)
        self.filtro_disp.currentIndexChanged.connect(self._buscar)

        iconos_crud(self.btn_nuevo, self.btn_editar, self.btn_eliminar,
                    self.btn_excel, self.btn_pdf, self.btn_buscar, self.btn_limpiar)

    def _aplicar_tema(self):
        t = obtener_tema()
        self.setStyleSheet(f"background-color: {t['fondo']}; color: {t['texto']};")
        self.titulo.setStyleSheet(f"font-size: 22px; font-weight: bold; color: {t['titulo']};")
        self.lbl_total.setStyleSheet(f"font-size: 11px; color: {t['texto_suave']};")
        self.btn_nuevo.setStyleSheet(estilo_boton(t["btn_primario"],  t["btn_primario_txt"]))
        self.btn_editar.setStyleSheet(estilo_boton(t["btn_editar"],   t["btn_editar_txt"]))
        self.btn_eliminar.setStyleSheet(estilo_boton(t["btn_peligro"],t["btn_peligro_txt"]))
        self.btn_excel.setStyleSheet(estilo_boton(t["btn_excel"],     t["btn_excel_txt"]))
        self.btn_pdf.setStyleSheet(estilo_boton(t["btn_pdf"],         t["btn_pdf_txt"]))
        self.btn_buscar.setStyleSheet(estilo_boton(t["btn_buscar"],   t["btn_buscar_txt"]))
        self.btn_limpiar.setStyleSheet(estilo_boton(t["btn_limpiar"], t["btn_limpiar_txt"]))
        self.campo_busqueda.setStyleSheet(estilo_input(t))
        self.filtro_nivel.setStyleSheet(estilo_input(t))
        self.filtro_disp.setStyleSheet(estilo_input(t))
        self.tabla.setStyleSheet(estilo_tabla(t))

    def _cargar_datos(self, consultores=None):
        if consultores is None:
            consultores = self.controller.obtener_todos()
        self.tabla.setRowCount(len(consultores))
        for fila, c in enumerate(consultores):
            self.tabla.setItem(fila, 0, QTableWidgetItem(str(c.codigo_empleado or "")))
            self.tabla.setItem(fila, 1, QTableWidgetItem(str(c.nombres or "")))
            self.tabla.setItem(fila, 2, QTableWidgetItem(str(c.apellidos or "")))
            self.tabla.setItem(fila, 3, QTableWidgetItem(str(c.nivel or "")))
            self.tabla.setItem(fila, 4, QTableWidgetItem(str(c.especialidades or "")))
            self.tabla.setItem(fila, 5, QTableWidgetItem(str(c.tarifa_horaria or "")))
            self.tabla.setItem(fila, 6, QTableWidgetItem(str(c.idiomas or "")))
            self.tabla.setItem(fila, 7, QTableWidgetItem(str(c.disponibilidad or "")))
        self.tabla.resizeColumnsToContents()
        self.lbl_total.setText(f"Total: {len(consultores)} registros")

    def _fila_seleccionada(self):
        fila = self.tabla.currentRow()
        if fila == -1:
            return None
        return self.tabla.item(fila, 0).text()

    def _buscar(self):
        termino = self.campo_busqueda.text().strip()
        nivel   = self.filtro_nivel.currentText()
        disp    = self.filtro_disp.currentText()
        resultados = self.controller.buscar(
            termino=termino,
            nivel=None if nivel == "Todos los niveles" else nivel,
            disponibilidad=None if disp == "Toda disponibilidad" else disp,
        )
        self._cargar_datos(resultados)

    def _limpiar(self):
        self.campo_busqueda.clear()
        self.filtro_nivel.setCurrentIndex(0)
        self.filtro_disp.setCurrentIndex(0)
        self._cargar_datos()

    def _abrir_formulario_nuevo(self):
        from views.dialogs.consultor_form_dialog import FormularioConsultor
        dialogo = FormularioConsultor(self.controller)
        if dialogo.exec() == QDialog.DialogCode.Accepted:
            self._cargar_datos()

    def _abrir_formulario_editar(self):
        codigo = self._fila_seleccionada()
        if not codigo:
            QMessageBox.warning(self, "Aviso", "Selecciona un consultor para editar.")
            return
        consultor = self.controller.obtener_por_codigo(codigo)
        from views.dialogs.consultor_form_dialog import FormularioConsultor
        dialogo = FormularioConsultor(self.controller, consultor)
        if dialogo.exec() == QDialog.DialogCode.Accepted:
            self._cargar_datos()

    def _eliminar(self):
        codigo = self._fila_seleccionada()
        if not codigo:
            QMessageBox.warning(self, "Aviso", "Selecciona un consultor para eliminar.")
            return
        respuesta = QMessageBox.question(
            self, "Confirmar",
            f"Eliminar el consultor {codigo}?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if respuesta == QMessageBox.StandardButton.Yes:
            try:
                self.controller.eliminar(codigo)
                self._cargar_datos()
                QMessageBox.information(self, "Exito", "Consultor eliminado correctamente.")
            except Exception as e:
                QMessageBox.critical(self, "Error", str(e))

    def _exportar_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            QMessageBox.critical(self, "Error", "Instala openpyxl:\npip install openpyxl")
            return
        consultores = self.controller.obtener_todos()
        ruta, _ = QFileDialog.getSaveFileName(self, "Guardar Excel", "consultores.xlsx", "Excel (*.xlsx)")
        if not ruta:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Consultores"
        cols = ["Codigo", "Nombres", "Apellidos", "Nivel", "Especialidades",
                "Tarifa/Hora", "Idiomas", "Disponibilidad"]
        for i, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=i, value=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2C3E50")
        for idx, c in enumerate(consultores, 2):
            ws.append([c.codigo_empleado, c.nombres, c.apellidos, c.nivel,
                       c.especialidades, str(c.tarifa_horaria or ""), c.idiomas, c.disponibilidad])
        wb.save(ruta)
        QMessageBox.information(self, "Exportado", f"Excel guardado en:\n{ruta}")

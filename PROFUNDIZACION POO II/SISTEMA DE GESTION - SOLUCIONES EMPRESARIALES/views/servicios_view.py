from datetime import date

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QTableWidget,
    QTableWidgetItem, QDialog, QLineEdit, QComboBox,
    QLabel, QMessageBox, QFileDialog, QDoubleSpinBox, QTextEdit, QFormLayout
)
from PyQt6.QtCore import Qt

from ui.themes import obtener_tema, estilo_boton, estilo_input, estilo_tabla
from ui.iconos.iconos import iconos_crud, iconos_formulario, favicon


class ServiciosView(QWidget):
    """
    Vista del modulo de Servicios.
    Delega logica al ServicioController.
    """

    def __init__(self, controller):
        super().__init__()
        self.controller = controller
        self._construir_ui()
        self._aplicar_tema()
        self._cargar_datos()

    def _construir_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(16, 16, 16, 16)
        self.setLayout(layout)

        self.titulo = QLabel("Servicios")
        layout.addWidget(self.titulo)

        barra = QHBoxLayout()
        self.btn_nuevo    = QPushButton("  Nuevo")
        self.btn_editar   = QPushButton("  Editar")
        self.btn_eliminar = QPushButton("  Eliminar")
        self.btn_excel    = QPushButton("  Exportar Excel")
        self.btn_pdf      = QPushButton("  Exportar PDF")
        for btn in [self.btn_nuevo, self.btn_editar, self.btn_eliminar,
                    self.btn_excel, self.btn_pdf]:
            btn.setMinimumHeight(34)
        barra.addWidget(self.btn_nuevo)
        barra.addWidget(self.btn_editar)
        barra.addWidget(self.btn_eliminar)
        barra.addSpacing(20)
        barra.addWidget(self.btn_excel)
        barra.addWidget(self.btn_pdf)
        barra.addStretch()
        layout.addLayout(barra)

        filtros = QHBoxLayout()
        self.campo_busqueda = QLineEdit()
        self.campo_busqueda.setPlaceholderText("Buscar por nombre o descripcion...")
        self.campo_busqueda.setMinimumHeight(32)
        self.filtro_tipo = QComboBox()
        self.filtro_tipo.addItems(["Todos los tipos", "consultoria", "implementacion", "soporte", "capacitacion"])
        self.filtro_tipo.setMinimumHeight(32)
        self.filtro_estado = QComboBox()
        self.filtro_estado.addItems(["Todos los estados", "activo", "inactivo"])
        self.filtro_estado.setMinimumHeight(32)
        self.btn_buscar  = QPushButton("  Buscar")
        self.btn_limpiar = QPushButton("  Limpiar")
        self.btn_buscar.setMinimumHeight(32)
        self.btn_limpiar.setMinimumHeight(32)
        filtros.addWidget(self.campo_busqueda, 3)
        filtros.addWidget(self.filtro_tipo, 1)
        filtros.addWidget(self.filtro_estado, 1)
        filtros.addWidget(self.btn_buscar)
        filtros.addWidget(self.btn_limpiar)
        layout.addLayout(filtros)

        self.tabla = QTableWidget()
        columnas = ["Codigo", "Nombre", "Tipo", "Descripcion", "Precio Base", "Unidad Cobro", "Estado"]
        self.tabla.setColumnCount(len(columnas))
        self.tabla.setHorizontalHeaderLabels(columnas)
        self.tabla.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.tabla.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.tabla.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.tabla.horizontalHeader().setStretchLastSection(True)
        self.tabla.verticalHeader().setVisible(False)
        self.tabla.setAlternatingRowColors(True)
        layout.addWidget(self.tabla)

        self.lbl_total = QLabel("Total: 0 registros")
        layout.addWidget(self.lbl_total)

        self.btn_nuevo.clicked.connect(self._abrir_formulario_nuevo)
        self.btn_editar.clicked.connect(self._abrir_formulario_editar)
        self.btn_eliminar.clicked.connect(self._eliminar)
        self.btn_excel.clicked.connect(self._exportar_excel)
        self.btn_buscar.clicked.connect(self._buscar)
        self.btn_limpiar.clicked.connect(self._limpiar)
        self.campo_busqueda.returnPressed.connect(self._buscar)
        self.filtro_tipo.currentIndexChanged.connect(self._buscar)
        self.filtro_estado.currentIndexChanged.connect(self._buscar)

        iconos_crud(self.btn_nuevo, self.btn_editar, self.btn_eliminar,
                    self.btn_excel, self.btn_pdf, self.btn_buscar, self.btn_limpiar)

    def _aplicar_tema(self):
        t = obtener_tema()
        self.setStyleSheet(f"background-color: {t['fondo']}; color: {t['texto']};")
        self.titulo.setStyleSheet(f"font-size: 22px; font-weight: bold; color: {t['titulo']};")
        self.lbl_total.setStyleSheet(f"font-size: 11px; color: {t['texto_suave']};")
        self.btn_nuevo.setStyleSheet(estilo_boton(t["btn_primario"],  t["btn_primario_txt"]))
        self.btn_editar.setStyleSheet(estilo_boton(t["btn_editar"],   t["btn_editar_txt"]))
        self.btn_eliminar.setStyleSheet(estilo_boton(t["btn_peligro"],t["btn_peligro_txt"]))
        self.btn_excel.setStyleSheet(estilo_boton(t["btn_excel"],     t["btn_excel_txt"]))
        self.btn_pdf.setStyleSheet(estilo_boton(t["btn_pdf"],         t["btn_pdf_txt"]))
        self.btn_buscar.setStyleSheet(estilo_boton(t["btn_buscar"],   t["btn_buscar_txt"]))
        self.btn_limpiar.setStyleSheet(estilo_boton(t["btn_limpiar"], t["btn_limpiar_txt"]))
        self.campo_busqueda.setStyleSheet(estilo_input(t))
        self.filtro_tipo.setStyleSheet(estilo_input(t))
        self.filtro_estado.setStyleSheet(estilo_input(t))
        self.tabla.setStyleSheet(estilo_tabla(t))

    def _cargar_datos(self, servicios=None):
        if servicios is None:
            servicios = self.controller.obtener_todos()
        self.tabla.setRowCount(len(servicios))
        for fila, s in enumerate(servicios):
            self.tabla.setItem(fila, 0, QTableWidgetItem(str(s.codigo or "")))
            self.tabla.setItem(fila, 1, QTableWidgetItem(str(s.nombre or "")))
            self.tabla.setItem(fila, 2, QTableWidgetItem(str(s.tipo or "")))
            self.tabla.setItem(fila, 3, QTableWidgetItem(str(s.descripcion or "")))
            self.tabla.setItem(fila, 4, QTableWidgetItem(str(s.precio_base or "")))
            self.tabla.setItem(fila, 5, QTableWidgetItem(str(s.unidad_cobro or "")))
            self.tabla.setItem(fila, 6, QTableWidgetItem(str(s.estado or "")))
        self.tabla.resizeColumnsToContents()
        self.lbl_total.setText(f"Total: {len(servicios)} registros")

    def _fila_seleccionada(self):
        fila = self.tabla.currentRow()
        if fila == -1:
            return None
        return self.tabla.item(fila, 0).text()

    def _buscar(self):
        termino = self.campo_busqueda.text().strip()
        tipo    = self.filtro_tipo.currentText()
        estado  = self.filtro_estado.currentText()
        resultados = self.controller.buscar(
            termino=termino,
            tipo=None if tipo == "Todos los tipos" else tipo,
            estado=None if estado == "Todos los estados" else estado,
        )
        self._cargar_datos(resultados)

    def _limpiar(self):
        self.campo_busqueda.clear()
        self.filtro_tipo.setCurrentIndex(0)
        self.filtro_estado.setCurrentIndex(0)
        self._cargar_datos()

    def _abrir_formulario_nuevo(self):
        from views.dialogs.servicio_form_dialog import FormularioServicio
        dialogo = FormularioServicio(self.controller)
        if dialogo.exec() == QDialog.DialogCode.Accepted:
            self._cargar_datos()

    def _abrir_formulario_editar(self):
        codigo = self._fila_seleccionada()
        if not codigo:
            QMessageBox.warning(self, "Aviso", "Selecciona un servicio para editar.")
            return
        servicio = self.controller.obtener_por_codigo(codigo)
        from views.dialogs.servicio_form_dialog import FormularioServicio
        dialogo = FormularioServicio(self.controller, servicio)
        if dialogo.exec() == QDialog.DialogCode.Accepted:
            self._cargar_datos()

    def _eliminar(self):
        codigo = self._fila_seleccionada()
        if not codigo:
            QMessageBox.warning(self, "Aviso", "Selecciona un servicio para eliminar.")
            return
        respuesta = QMessageBox.question(
            self, "Confirmar",
            f"Eliminar el servicio {codigo}?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if respuesta == QMessageBox.StandardButton.Yes:
            try:
                self.controller.eliminar(codigo)
                self._cargar_datos()
                QMessageBox.information(self, "Exito", "Servicio eliminado correctamente.")
            except Exception as e:
                QMessageBox.critical(self, "Error", str(e))

    def _exportar_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            QMessageBox.critical(self, "Error", "Instala openpyxl:\npip install openpyxl")
            return
        servicios = self.controller.obtener_todos()
        ruta, _ = QFileDialog.getSaveFileName(self, "Guardar Excel", "servicios.xlsx", "Excel (*.xlsx)")
        if not ruta:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Servicios"
        cols = ["Codigo", "Nombre", "Tipo", "Descripcion", "Precio Base", "Unidad Cobro", "Estado"]
        for i, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=i, value=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2C3E50")
        for s in servicios:
            ws.append([s.codigo, s.nombre, s.tipo, s.descripcion,
                       str(s.precio_base or ""), s.unidad_cobro, s.estado])
        wb.save(ruta)
        QMessageBox.information(self, "Exportado", f"Excel guardado en:\n{ruta}")
